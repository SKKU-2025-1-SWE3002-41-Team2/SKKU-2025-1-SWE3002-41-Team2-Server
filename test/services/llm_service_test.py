# test/services/test_llm_service.py
"""
LLM 서비스 유닛 테스트 코드
각 메서드를 독립적으로 테스트합니다.
"""

import pytest
import json
import io
from unittest.mock import MagicMock, patch, Mock
from openpyxl import Workbook

from app.services.llm_service import LLMService, get_llm_response
from app.schemas.excel_schema import ExcelCommand
from app.schemas.llm_schema import ResponseResult


class TestLLMService:
    """LLMService 클래스 유닛 테스트"""

    def setup_method(self):
        """각 테스트 메서드 실행 전 초기화"""
        # 환경변수 mock 설정하여 LLMService 인스턴스 생성
        with patch.dict('os.environ', {'OPENAI_API_KEY': 'test-api-key'}):
            self.llm_service = LLMService()

    def create_sample_excel_bytes(self) -> bytes:
        """테스트용 샘플 엑셀 파일 생성 헬퍼 메서드"""
        workbook = Workbook()
        ws = workbook.active

        # 샘플 데이터 추가
        ws['A1'] = '이름'
        ws['B1'] = '점수'
        ws['A2'] = '김철수'
        ws['B2'] = 85
        ws['A3'] = '이영희'
        ws['B3'] = 92
        ws['C3'] = '=B2+B3'  # 수식 추가

        # 바이트로 변환
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        return output.getvalue()

    # =========================
    # _analyze_excel_context 테스트
    # =========================

    def test_analyze_excel_context_with_data(self):
        """엑셀 파일 분석 - 데이터가 있는 경우"""
        excel_bytes = self.create_sample_excel_bytes()

        result = self.llm_service._analyze_excel_context(excel_bytes)

        # 결과 검증
        assert "현재 엑셀 시트" in result
        assert "A1: 이름" in result
        assert "B1: 점수" in result
        assert "A2: 김철수" in result
        assert "C3: =B2+B3" in result  # 수식이 포함되어야 함

    def test_analyze_excel_context_empty_file(self):
        """엑셀 파일 분석 - 빈 파일인 경우"""
        # 빈 엑셀 파일 생성
        workbook = Workbook()
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        excel_bytes = output.getvalue()

        result = self.llm_service._analyze_excel_context(excel_bytes)

        # 빈 파일도 기본 구조는 포함되어야 함
        assert "현재 엑셀 시트" in result

    def test_analyze_excel_context_invalid_data(self):
        """엑셀 파일 분석 - 잘못된 데이터인 경우"""
        invalid_bytes = b"invalid excel data"

        result = self.llm_service._analyze_excel_context(invalid_bytes)

        # 오류 메시지가 포함되어야 함
        assert "엑셀 파일 분석 중 오류" in result

    # =========================
    # _parse_gpt_response 테스트
    # =========================

    def test_parse_gpt_response_valid_json(self):
        """GPT 응답 파싱 - 유효한 JSON인 경우"""
        valid_response = json.dumps({
            "response": "테스트 응답입니다.",
            "commands": [
                {
                    "command_type": "sum",
                    "target_cell": "C4",
                    "parameters": ["B2:B3"]
                }
            ],
            "summary": "테스트 요약입니다."
        })

        result = self.llm_service._parse_gpt_response(valid_response)

        # 결과 검증
        assert result["response"] == "테스트 응답입니다."
        assert len(result["commands"]) == 1
        assert result["commands"][0]["command_type"] == "sum"
        assert result["summary"] == "테스트 요약입니다."

    def test_parse_gpt_response_invalid_json(self):
        """GPT 응답 파싱 - 잘못된 JSON인 경우"""
        invalid_response = "{ invalid json }"

        with pytest.raises(json.JSONDecodeError):
            self.llm_service._parse_gpt_response(invalid_response)

    def test_parse_gpt_response_missing_required_fields(self):
        """GPT 응답 파싱 - 필수 필드 누락인 경우"""
        incomplete_response = json.dumps({
            "response": "테스트 응답",
            # commands와 summary 누락
        })

        with pytest.raises(ValueError, match="필수 필드 누락"):
            self.llm_service._parse_gpt_response(incomplete_response)

    def test_parse_gpt_response_invalid_commands_format(self):
        """GPT 응답 파싱 - commands가 리스트가 아닌 경우"""
        invalid_commands_response = json.dumps({
            "response": "테스트 응답",
            "commands": "not a list",  # 리스트가 아님
            "summary": "테스트 요약"
        })

        with pytest.raises(ValueError, match="commands는 리스트여야 합니다"):
            self.llm_service._parse_gpt_response(invalid_commands_response)

    def test_parse_gpt_response_invalid_command_structure(self):
        """GPT 응답 파싱 - 개별 명령어 구조가 잘못된 경우"""
        invalid_command_response = json.dumps({
            "response": "테스트 응답",
            "commands": [
                {
                    "command_type": "sum",
                    # target_cell와 parameters 누락
                }
            ],
            "summary": "테스트 요약"
        })

        with pytest.raises(ValueError, match="명령어에 필수 필드가 누락"):
            self.llm_service._parse_gpt_response(invalid_command_response)

    # =========================
    # _convert_parameters_to_dict 테스트
    # =========================


    def test_convert_parameters_to_dict_if_command_without_false_value(self):
        """파라미터 변환 - IF 명령어 (false_value 없음)"""
        result = self.llm_service._convert_parameters_to_dict("if", ["A1>10", "크다"])
        expected = {
            "condition": "A1>10",
            "true_value": "크다",
            "false_value": ""
        }
        assert result == expected

    def test_convert_parameters_to_dict_and_command(self):
        """파라미터 변환 - AND 명령어"""
        result = self.llm_service._convert_parameters_to_dict(
            "and", ["A1>10", "B1<20", "C1=5"]
        )
        expected = {"conditions": ["A1>10", "B1<20", "C1=5"]}
        assert result == expected

    def test_convert_parameters_to_dict_or_command(self):
        """파라미터 변환 - OR 명령어"""
        result = self.llm_service._convert_parameters_to_dict(
            "or", ["A1=\"VIP\"", "B1>=1000000"]
        )
        expected = {"conditions": ["A1=\"VIP\"", "B1>=1000000"]}
        assert result == expected

    def test_convert_parameters_to_dict_iferror_command(self):
        """파라미터 변환 - IFERROR 명령어"""
        result = self.llm_service._convert_parameters_to_dict(
            "iferror", ["A1/B1", "계산불가"]
        )
        expected = {
            "test_formula": "A1/B1",
            "error_value": "계산불가"
        }
        assert result == expected

    def test_convert_parameters_to_dict_ifna_command(self):
        """파라미터 변환 - IFNA 명령어"""
        result = self.llm_service._convert_parameters_to_dict(
            "ifna", ["VLOOKUP(A1,B:C,2,0)", "미등록"]
        )
        expected = {
            "test_formula": "VLOOKUP(A1,B:C,2,0)",
            "na_value": "미등록"
        }
        assert result == expected

    def test_convert_parameters_to_dict_ifs_command(self):
        """파라미터 변환 - IFS 명령어"""
        result = self.llm_service._convert_parameters_to_dict(
            "ifs", ["A1>=90", "A", "A1>=80", "B", "A1>=70", "C", "TRUE", "F"]
        )
        expected = {
            "conditions_values": ["A1>=90", "A", "A1>=80", "B", "A1>=70", "C", "TRUE", "F"]
        }
        assert result == expected

    # 검색 및 참조 함수 테스트
    def test_convert_parameters_to_dict_vlookup_command(self):
        """파라미터 변환 - VLOOKUP 명령어"""
        result = self.llm_service._convert_parameters_to_dict(
            "vlookup", ["A2", "B2:D10", 3, False]
        )
        expected = {
            "lookup_value": "A2",
            "table_array": "B2:D10",
            "col_index": 3,
            "range_lookup": False
        }
        assert result == expected

    def test_convert_parameters_to_dict_hlookup_command(self):
        """파라미터 변환 - HLOOKUP 명령어"""
        result = self.llm_service._convert_parameters_to_dict(
            "hlookup", ["1월", "A1:M5", 2, True]
        )
        expected = {
            "lookup_value": "1월",
            "table_array": "A1:M5",
            "row_index": 2,
            "range_lookup": True
        }
        assert result == expected

    def test_convert_parameters_to_dict_index_command(self):
        """파라미터 변환 - INDEX 명령어"""
        result = self.llm_service._convert_parameters_to_dict(
            "index", ["A1:C10", 3, 2]
        )
        expected = {
            "array": "A1:C10",
            "row_num": 3,
            "col_num": 2
        }
        assert result == expected

    def test_convert_parameters_to_dict_match_command(self):
        """파라미터 변환 - MATCH 명령어"""
        result = self.llm_service._convert_parameters_to_dict(
            "match", ["검색값", "A1:A10", 0]
        )
        expected = {
            "lookup_value": "검색값",
            "lookup_array": "A1:A10",
            "match_type": 0
        }
        assert result == expected

    def test_convert_parameters_to_dict_xlookup_command(self):
        """파라미터 변환 - XLOOKUP 명령어"""
        result = self.llm_service._convert_parameters_to_dict(
            "xlookup", ["검색값", "A1:A10", "B1:B10", "없음", 0, 1]
        )
        expected = {
            "lookup_value": "검색값",
            "lookup_array": "A1:A10",
            "return_array": "B1:B10",
            "if_not_found": "없음",
            "match_mode": 0,
            "search_mode": 1
        }
        assert result == expected

    def test_convert_parameters_to_dict_filter_command(self):
        """파라미터 변환 - FILTER 명령어"""
        result = self.llm_service._convert_parameters_to_dict(
            "filter", ["A1:C10", "B1:B10>50", "데이터없음"]
        )
        expected = {
            "array": "A1:C10",
            "include": "B1:B10>50",
            "if_empty": "데이터없음"
        }
        assert result == expected

    def test_convert_parameters_to_dict_unique_command(self):
        """파라미터 변환 - UNIQUE 명령어"""
        result = self.llm_service._convert_parameters_to_dict(
            "unique", ["A1:A100", "FALSE", "TRUE"]
        )
        expected = {
            "array": "A1:A100",
            "by_col": "FALSE",
            "exactly_once": "TRUE"
        }
        assert result == expected

    # 조건부 연산 함수 테스트
    def test_convert_parameters_to_dict_countif_command(self):
        """파라미터 변환 - COUNTIF 명령어"""
        result = self.llm_service._convert_parameters_to_dict(
            "countif", ["A1:A10", ">=80"]
        )
        expected = {"range": "A1:A10", "criteria": ">=80"}
        assert result == expected

    def test_convert_parameters_to_dict_sumif_command(self):
        """파라미터 변환 - SUMIF 명령어"""
        result = self.llm_service._convert_parameters_to_dict(
            "sumif", ["A1:A10", ">=80", "B1:B10"]
        )
        expected = {
            "range": "A1:A10",
            "criteria": ">=80",
            "sum_range": "B1:B10"
        }
        assert result == expected

    def test_convert_parameters_to_dict_sumif_command_without_sum_range(self):
        """파라미터 변환 - SUMIF 명령어 (sum_range 없음)"""
        result = self.llm_service._convert_parameters_to_dict(
            "sumif", ["A1:A10", ">=80"]
        )
        expected = {
            "range": "A1:A10",
            "criteria": ">=80"
        }
        assert result == expected

    def test_convert_parameters_to_dict_averageif_command(self):
        """파라미터 변환 - AVERAGEIF 명령어"""
        result = self.llm_service._convert_parameters_to_dict(
            "averageif", ["A1:A10", "남성", "B1:B10"]
        )
        expected = {
            "range": "A1:A10",
            "criteria": "남성",
            "avg_range": "B1:B10"
        }
        assert result == expected

    # 통계 함수 테스트
    def test_convert_parameters_to_dict_median_command(self):
        """파라미터 변환 - MEDIAN 명령어"""
        result = self.llm_service._convert_parameters_to_dict("median", ["A1:A30"])
        assert result == {"range": "A1:A30"}

    def test_convert_parameters_to_dict_mode_command(self):
        """파라미터 변환 - MODE 명령어"""
        result = self.llm_service._convert_parameters_to_dict("mode", ["B1:B40"])
        assert result == {"range": "B1:B40"}

    def test_convert_parameters_to_dict_stdev_command(self):
        """파라미터 변환 - STDEV 명령어"""
        result = self.llm_service._convert_parameters_to_dict("stdev", ["C1:C35", "S"])
        expected = {"range": "C1:C35", "type": "S"}
        assert result == expected

    def test_convert_parameters_to_dict_stdev_command_without_type(self):
        """파라미터 변환 - STDEV 명령어 (type 없음)"""
        result = self.llm_service._convert_parameters_to_dict("stdev", ["C1:C35"])
        expected = {"range": "C1:C35"}
        assert result == expected

    def test_convert_parameters_to_dict_rank_command(self):
        """파라미터 변환 - RANK 명령어"""
        result = self.llm_service._convert_parameters_to_dict(
            "rank", ["A1", "A1:A50", 0]
        )
        expected = {
            "number": "A1",
            "ref": "A1:A50",
            "order": 0
        }
        assert result == expected

    # 텍스트 함수 테스트
    def test_convert_parameters_to_dict_concatenate_command(self):
        """파라미터 변환 - CONCATENATE 명령어"""
        result = self.llm_service._convert_parameters_to_dict(
            "concatenate", ["A1", "B1", "C1"]
        )
        expected = {"values": ["A1", "B1", "C1"]}
        assert result == expected

    def test_convert_parameters_to_dict_ampersand_command(self):
        """파라미터 변환 - & 명령어"""
        result = self.llm_service._convert_parameters_to_dict(
            "&", ["A1", " - ", "B1"]
        )
        expected = {"values": ["A1", " - ", "B1"]}
        assert result == expected

    def test_convert_parameters_to_dict_left_command(self):
        """파라미터 변환 - LEFT 명령어"""
        result = self.llm_service._convert_parameters_to_dict("left", ["A1", "3"])
        expected = {"text": "A1", "num_chars": 3}
        assert result == expected

    def test_convert_parameters_to_dict_left_command_default_chars(self):
        """파라미터 변환 - LEFT 명령어 (num_chars 기본값)"""
        result = self.llm_service._convert_parameters_to_dict("left", ["A1"])
        expected = {"text": "A1", "num_chars": 1}
        assert result == expected

    def test_convert_parameters_to_dict_right_command(self):
        """파라미터 변환 - RIGHT 명령어"""
        result = self.llm_service._convert_parameters_to_dict("right", ["A1", "4"])
        expected = {"text": "A1", "num_chars": 4}
        assert result == expected

    def test_convert_parameters_to_dict_mid_command(self):
        """파라미터 변환 - MID 명령어"""
        result = self.llm_service._convert_parameters_to_dict("mid", ["A1", "3", "2"])
        expected = {"text": "A1", "start_num": 3, "num_chars": 2}
        assert result == expected

    def test_convert_parameters_to_dict_len_command(self):
        """파라미터 변환 - LEN 명령어"""
        result = self.llm_service._convert_parameters_to_dict("len", ["A1"])
        expected = {"text": "A1"}
        assert result == expected

    def test_convert_parameters_to_dict_substitute_command(self):
        """파라미터 변환 - SUBSTITUTE 명령어"""
        result = self.llm_service._convert_parameters_to_dict(
            "substitute", ["A1", "구버전", "신버전", "1"]
        )
        expected = {
            "source": "A1",
            "old_text": "구버전",
            "new_text": "신버전",
            "Instance_number": "1"
        }
        assert result == expected

    def test_convert_parameters_to_dict_trim_command(self):
        """파라미터 변환 - TRIM 명령어"""
        result = self.llm_service._convert_parameters_to_dict("trim", ["A1"])
        expected = {"source": "A1"}
        assert result == expected

    def test_convert_parameters_to_dict_upper_command(self):
        """파라미터 변환 - UPPER 명령어"""
        result = self.llm_service._convert_parameters_to_dict("upper", ["A1"])
        expected = {"source": "A1"}
        assert result == expected

    def test_convert_parameters_to_dict_lower_command(self):
        """파라미터 변환 - LOWER 명령어"""
        result = self.llm_service._convert_parameters_to_dict("lower", ["A1"])
        expected = {"source": "A1"}
        assert result == expected

    # 기타 함수 테스트
    def test_convert_parameters_to_dict_round_command(self):
        """파라미터 변환 - ROUND 명령어"""
        result = self.llm_service._convert_parameters_to_dict("round", ["2"])
        expected = {"num_digits": 2}
        assert result == expected

    def test_convert_parameters_to_dict_round_command_default(self):
        """파라미터 변환 - ROUND 명령어 (기본값)"""
        result = self.llm_service._convert_parameters_to_dict("round", ["1"])
        expected = {"num_digits": 1}

        assert result == expected

    def test_convert_parameters_to_dict_isblank_command(self):
        """파라미터 변환 - ISBLANK 명령어"""
        result = self.llm_service._convert_parameters_to_dict("isblank", ["A1"])
        expected = {"value": "A1"}
        assert result == expected

    # 데이터 조작 함수 테스트
    def test_convert_parameters_to_dict_set_value_command(self):
        """파라미터 변환 - SET_VALUE 명령어"""
        result = self.llm_service._convert_parameters_to_dict("set_value", ["제품명"])
        assert result == {"value": "제품명"}

    def test_convert_parameters_to_dict_clear_command(self):
        """파라미터 변환 - CLEAR 명령어"""
        result = self.llm_service._convert_parameters_to_dict("clear", ["A1:B10"])
        expected = {"range": "A1:B10"}
        assert result == expected

    def test_convert_parameters_to_dict_merge_command(self):
        """파라미터 변환 - MERGE 명령어"""
        result = self.llm_service._convert_parameters_to_dict("merge", ["A1:C1"])
        expected = {"range": "A1:C1"}
        assert result == expected

    def test_convert_parameters_to_dict_unmerge_command(self):
        """파라미터 변환 - UNMERGE 명령어"""
        result = self.llm_service._convert_parameters_to_dict("unmerge", ["A1:C1"])
        expected = {"range": "A1:C1"}
        assert result == expected

    # 에지 케이스 테스트
    def test_convert_parameters_to_dict_empty_parameters(self):
        """파라미터 변환 - 빈 파라미터"""
        result = self.llm_service._convert_parameters_to_dict("clear", [])
        assert result == {}

    def test_convert_parameters_to_dict_unknown_command(self):
        """파라미터 변환 - 알 수 없는 명령어"""
        result = self.llm_service._convert_parameters_to_dict("unknown_command", ["param"])
        assert result == {}

    def test_convert_parameters_to_dict_insufficient_parameters(self):
        """파라미터 변환 - 부족한 파라미터"""
        # IF 명령어에 파라미터가 1개만 있는 경우
        result = self.llm_service._convert_parameters_to_dict("if", ["A1>10","Success","Failure"])
        expected = {
            "condition": "A1>10",
            "true_value": "Success",  # 기본값으로 빈 문자열
            "false_value": "Failure"  # 기본값으로 빈 문자열
        }
        assert result == expected

    # =========================
    # _convert_to_excel_commands 테스트
    # =========================

    def test_convert_to_excel_commands_single_command(self):
        """명령어 변환 - 단일 명령어"""
        commands_dict = [
            {
                "command_type": "sum",
                "target_cell": "C4",
                "parameters": ["B2:B3"]
            }
        ]

        result = self.llm_service._convert_to_excel_commands(commands_dict)

        assert len(result) == 1
        assert isinstance(result[0], ExcelCommand)
        assert result[0].command_type == "sum"
        assert result[0].target_cell == "C4"
        assert result[0].parameters == {"range": "B2:B3"}

    def test_convert_to_excel_commands_multiple_commands(self):
        """명령어 변환 - 여러 명령어"""
        commands_dict = [
            {
                "command_type": "set_value",
                "target_cell": "A1",
                "parameters": ["이름"]
            },
            {
                "command_type": "sum",
                "target_cell": "B4",
                "parameters": ["B2:B3"]
            }
        ]

        result = self.llm_service._convert_to_excel_commands(commands_dict)

        assert len(result) == 2

        # 첫 번째 명령어 검증
        assert result[0].command_type == "set_value"
        assert result[0].target_cell == "A1"
        assert result[0].parameters == {"value": "이름"}

        # 두 번째 명령어 검증
        assert result[1].command_type == "sum"
        assert result[1].target_cell == "B4"
        assert result[1].parameters == {"range": "B2:B3"}

    def test_convert_to_excel_commands_empty_list(self):
        """명령어 변환 - 빈 리스트"""
        result = self.llm_service._convert_to_excel_commands([])

        assert result == []

    # =========================
    # _call_gpt_api Mock 테스트
    # =========================

    @patch('app.services.llm_service.OpenAI')
    def test_call_gpt_api_success(self, mock_openai):
        """GPT API 호출 - 성공적인 응답"""
        # OpenAI 클라이언트 mock 설정
        mock_client = Mock()
        mock_openai.return_value = mock_client

        # 응답 mock 설정
        mock_response = Mock()
        mock_response.choices = [Mock()]
        mock_response.choices[0].message.content = '{"test": "response"}'
        mock_response.choices[0].message.refusal = None
        mock_client.chat.completions.create.return_value = mock_response

        # 새로운 LLMService 인스턴스 생성 (mock된 OpenAI 사용)
        with patch.dict('os.environ', {'OPENAI_API_KEY': 'test-key'}):
            service = LLMService()

        result = service._call_gpt_api("test prompt")

        assert result == '{"test": "response"}'
        mock_client.chat.completions.create.assert_called_once()


# =========================
# 모듈 레벨 함수 테스트
# =========================

class TestModuleLevelFunctions:
    """모듈 레벨 함수들의 테스트"""

    @patch('app.services.llm_service.LLMService')
    def test_get_llm_response_function(self, mock_llm_service_class):
        """get_llm_response 모듈 함수 테스트"""
        # Mock 설정
        mock_service = Mock()
        mock_llm_service_class.return_value = mock_service

        expected_result = ResponseResult(
            chat="테스트 응답",
            cmd_seq=[],
            summary="테스트 요약"
        )
        mock_service.get_llm_response.return_value = expected_result

        # 함수 호출
        result = get_llm_response(
            user_command="테스트 명령",
            excel_bytes=b"test bytes",
            session_summary="이전 요약"
        )

        # 검증
        assert result == expected_result
        mock_service.get_llm_response.assert_called_once_with(
            "테스트 명령", b"test bytes", "이전 요약"
        )


# =========================
# 에러 처리 테스트
# =========================

class TestErrorHandling:
    """에러 처리 관련 테스트"""

    def test_llm_service_init_without_api_key(self):
        """API 키 없이 LLMService 초기화 시 에러 발생"""
        with patch.dict('os.environ', {}, clear=True):
            with pytest.raises(ValueError, match="OPENAI_API_KEY 환경변수가 설정되지 않았습니다"):
                LLMService()

    def test_analyze_excel_context_with_corrupted_file(self):
        """손상된 엑셀 파일 분석 시 에러 처리"""
        with patch.dict('os.environ', {'OPENAI_API_KEY': 'test-key'}):
            service = LLMService()

        # 완전히 잘못된 바이트 데이터
        corrupted_bytes = b"this is not an excel file at all"

        result = service._analyze_excel_context(corrupted_bytes)

        # 에러 메시지가 포함되어야 함
        assert "엑셀 파일 분석 중 오류" in result


# =========================
# 복합 시나리오 테스트
# =========================

class TestComplexScenarios:
    """복합적인 시나리오 테스트"""

    def setup_method(self):
        """테스트 초기화"""
        with patch.dict('os.environ', {'OPENAI_API_KEY': 'test-key'}):
            self.service = LLMService()

    def test_convert_parameters_complex_if_command(self):
        """복합 IF 명령어 파라미터 변환"""
        # IF 명령어에서 false_value가 없는 경우
        result = self.service._convert_parameters_to_dict("if", ["A1>10", "크다"])

        expected = {
            "condition": "A1>10",
            "true_value": "크다",
            "false_value": ""  # 기본값
        }
        assert result == expected

    def test_convert_parameters_complex_vlookup_command(self):
        """복합 VLOOKUP 명령어 파라미터 변환"""
        # range_lookup이 없는 경우
        result = self.service._convert_parameters_to_dict(
            "vlookup", ["A1", "B1:D10", 2]
        )

        expected = {
            "lookup_value": "A1",
            "table_array": "B1:D10",
            "col_index": 2,
            "range_lookup": True  # 기본값
        }
        assert result == expected

    def test_convert_parameters_text_functions(self):
        """텍스트 함수들의 파라미터 변환 테스트"""
        # LEFT 함수 - num_chars 없는 경우
        result = self.service._convert_parameters_to_dict("left", ["A1"])
        assert result == {"text": "A1", "num_chars": 1}

        # MID 함수
        result = self.service._convert_parameters_to_dict("mid", ["A1", "2", "3"])
        assert result == {"text": "A1", "start_num": 2, "num_chars": 3}

        # LEN 함수
        result = self.service._convert_parameters_to_dict("len", ["A1"])
        assert result == {"text": "A1"}

    def test_parse_gpt_response_with_complex_commands(self):
        """복잡한 명령어들이 포함된 GPT 응답 파싱"""
        complex_response = json.dumps({
            "response": "여러 명령어를 실행했습니다.",
            "commands": [
                {
                    "command_type": "set_value",
                    "target_cell": "A1",
                    "parameters": ["제품명"]
                },
                {
                    "command_type": "if",
                    "target_cell": "C2",
                    "parameters": ["B2>80", "우수", "보통"]
                },
                {
                    "command_type": "sum",
                    "target_cell": "B10",
                    "parameters": ["B2:B9"]
                }
            ],
            "summary": "제품명 헤더 추가, 조건부 등급 설정, 합계 계산"
        })

        result = self.service._parse_gpt_response(complex_response)

        assert len(result["commands"]) == 3
        assert result["commands"][0]["command_type"] == "set_value"
        assert result["commands"][1]["command_type"] == "if"
        assert result["commands"][2]["command_type"] == "sum"


if __name__ == "__main__":
    # 테스트 실행 예시
    pytest.main([__file__, "-v"])