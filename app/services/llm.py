# app/services/llm.py
"""
LLM 서비스 모듈
OpenAI GPT를 사용하여 자연어 명령을 엑셀 명령어로 변환하는 서비스
"""
import json
import os
from typing import List, Dict, Any, Optional
from openai import OpenAI
import io

from openpyxl import load_workbook

from app.schemas.llm import LLMResultInternal, ResponseResult
from app.services.llm_prompt import (
    SYSTEM_PROMPT,
    RESPONSE_SCHEMA,
    create_user_prompt,
    create_excel_context
)

# 타입 힌트를 위한 임포트
from app.schemas.excel import ExcelCommand


class LLMService:
    """
    LLM(Large Language Model) 서비스 클래스
    사용자의 자연어 명령을 엑셀 명령어로 변환합니다.
    """

    def __init__(self):
        """
        LLMService 초기화
        환경변수에서 OpenAI API 키를 가져와 클라이언트를 생성합니다.
        """
        api_key = os.getenv("OPENAI_API_KEY")
        if not api_key:
            raise ValueError("OPENAI_API_KEY 환경변수가 설정되지 않았습니다.")

        self.client = OpenAI(api_key=api_key)

    def get_llm_response(
            self,
            user_command: str,
            excel_bytes: bytes,
            session_summary: Optional[str] = None
    ) -> ResponseResult:
        """
        사용자의 명령을 받아 LLM으로 처리하고 결과를 반환합니다.

        Args:
            user_command: 사용자가 입력한 자연어 명령
            excel_bytes: 현재 엑셀 파일의 바이트 데이터
            session_summary: 이전 대화 요약 (옵션)

        Returns:
            ResponseResult: LLM 응답 결과 (chat, cmd_seq, summary)
        """
        # 1. 엑셀 파일 분석하여 컨텍스트 생성
        excel_context = self._analyze_excel_context(excel_bytes)

        # 2. 사용자 프롬프트 생성
        user_prompt = create_user_prompt(
            summary=session_summary or "",
            user_command=user_command,
            excel_context=excel_context
        )

        # 3. GPT API 호출
        try:
            response = self._call_gpt_api(user_prompt)

            # 4. 응답 파싱 및 검증
            parsed_response = self._parse_gpt_response(response)

            # 5. ExcelCommand 객체 리스트로 변환
            excel_commands = self._convert_to_excel_commands(parsed_response["commands"])

            # 6. 결과 반환
            return ResponseResult(
                chat=parsed_response["response"],
                cmd_seq=excel_commands,  # ExcelCommand 객체 리스트를 그대로 반환
                summary=parsed_response["summary"]
            )

        except Exception as e:
            print(f"LLM 처리 중 오류 발생: {str(e)}")
            # 에러 발생시 기본 응답 반환
            return ResponseResult(
                chat="죄송합니다. 명령을 처리하는 중 오류가 발생했습니다. 다시 시도해주세요.",
                cmd_seq=[],
                summary=session_summary or ""
            )

    def _analyze_excel_context(self, excel_bytes: bytes) -> str:
        """
        엑셀 파일을 분석하여 GPT가 이해할 수 있는 텍스트로 변환합니다.

        Args:
            excel_bytes: 엑셀 파일의 바이트 데이터

        Returns:
            엑셀 파일의 현재 상태를 설명하는 텍스트
        """
        try:
            # 엑셀 파일 로드
            workbook = load_workbook(io.BytesIO(excel_bytes))
            ws = workbook.active

            # 데이터가 있는 범위 확인
            max_row = ws.max_row
            max_col = ws.max_column

            # 데이터 샘플 수집
            sample_data = []
            formula_cells = []

            # 최대 10x10 범위까지 샘플링
            for row in range(1, min(11, max_row + 1)):
                for col in range(1, min(11, max_col + 1)):
                    cell = ws.cell(row=row, column=col)
                    if cell.value is not None:
                        cell_ref = cell.coordinate

                        # 수식인지 확인
                        if isinstance(cell.value, str) and cell.value.startswith('='):
                            formula_cells.append(f"{cell_ref}: {cell.value}")
                        else:
                            sample_data.append(f"{cell_ref}: {cell.value}")

            # 컨텍스트 생성
            return create_excel_context(
                rows=max_row,
                cols=max_col,
                sample_data=sample_data[:20],  # 최대 20개 샘플
                formula_data=formula_cells
            )

        except Exception as e:
            return f"엑셀 파일 분석 중 오류: {str(e)}"

    def _call_gpt_api(self, user_prompt: str) -> str:
        """
        OpenAI GPT API를 호출합니다.

        Args:
            user_prompt: 사용자 프롬프트

        Returns:
            GPT의 응답 텍스트
        """
        # API 호출
        completion = self.client.responses.parse(
            model="gpt-4.1",  # 또는 "gpt-4o" 사용 가능
            input=  [
                {"role": "system", "content": SYSTEM_PROMPT},
                {"role": "user", "content": user_prompt}
            ],
            response_format=RESPONSE_SCHEMA,  # llm_prompt.py에서 가져온 스키마 사용
            temperature=0.7,
            #max_tokens=32768,  # 최대 토큰 수
        )

        response = completion.choices[0].message
        # If the model refuses to respond, you will get a refusal message
        if(response.refusal):
            print(response.refusal)

        #logging
        print(f"GPT 응답: {response.content}")
        # 응답 반환
        return completion.choices[0].message.content

    def _parse_gpt_response(self, response: str) -> Dict[str, Any]:
        """
        GPT의 응답을 파싱하여 딕셔너리로 변환합니다.

        Args:
            response: GPT의 JSON 형식 응답

        Returns:
            파싱된 응답 딕셔너리
        """
        try:
            # JSON 파싱
            parsed = json.loads(response)

            # 필수 필드 검증
            required_fields = ["response", "commands", "summary"]
            for field in required_fields:
                if field not in parsed:
                    raise ValueError(f"필수 필드 누락: {field}")

            # commands가 리스트인지 확인
            if not isinstance(parsed["commands"], list):
                raise ValueError("commands는 리스트여야 합니다")

            # 각 명령어 검증
            for cmd in parsed["commands"]:
                if not all(key in cmd for key in ["command_type", "target_range", "parameters"]):
                    raise ValueError("명령어에 필수 필드가 누락되었습니다")
                if not isinstance(cmd["parameters"], list):
                    raise ValueError("parameters는 배열이어야 합니다")

            return parsed

        except json.JSONDecodeError as e:
            print(f"JSON 파싱 오류: {str(e)}")
            print(f"응답 내용: {response}")
            raise
        except Exception as e:
            print(f"응답 파싱 중 오류: {str(e)}")
            raise

    def _convert_to_excel_commands(self, commands: List[Dict[str, Any]]) -> List[ExcelCommand]:
        """
        파싱된 명령어를 ExcelCommand 객체 리스트로 변환합니다.

        Args:
            commands: 파싱된 명령어 딕셔너리 리스트

        Returns:
            ExcelCommand 객체 리스트
        """
        excel_commands = []

        for cmd in commands:
            # parameters 배열을 딕셔너리로 변환
            # 명령어 타입에 따라 적절한 키-값 쌍으로 변환
            parameters_dict = self._convert_parameters_to_dict(
                cmd["command_type"],
                cmd["parameters"]
            )

            excel_command = ExcelCommand(
                command_type=cmd["command_type"],
                target_range=cmd["target_range"],
                parameters=parameters_dict
            )
            excel_commands.append(excel_command)

        return excel_commands

    def _convert_parameters_to_dict(self, command_type: str, parameters: List[Any]) -> Dict[str, Any]:
        """
        명령어 타입에 따라 parameters 배열을 적절한 딕셔너리로 변환합니다.

        Args:
            command_type: 명령어 타입
            parameters: 파라미터 배열

        Returns:
            파라미터 딕셔너리
        """
        # 파라미터가 없는 경우
        if not parameters:
            return {}

        # 명령어 타입별 변환 규칙
        if command_type in ["sum", "average", "count", "max", "min"]:
            # 수식 함수: 첫 번째 파라미터가 범위
            return {"range": parameters[0]} if parameters else {}

        elif command_type in ["font_color", "fill_color"]:
            # 색상 설정: 첫 번째 파라미터가 색상 값
            return {"color": parameters[0]} if parameters else {}

        elif command_type == "font_size":
            # 글자 크기: 첫 번째 파라미터가 크기
            return {"size": int(parameters[0])} if parameters else {}

        elif command_type == "font_name":
            # 글꼴: 첫 번째 파라미터가 글꼴 이름
            return {"name": parameters[0]} if parameters else {}

        elif command_type == "border":
            # 테두리: 스타일 지정 (기본값: thin)
            return {"style": parameters[0] if parameters else "thin"}

        elif command_type == "set_value":
            # 값 설정: 첫 번째 파라미터가 값
            return {"value": parameters[0]} if parameters else {}

        else:
            # 기타 명령어는 파라미터가 필요 없음
            return {}


# 모듈 레벨 함수로 export
def get_llm_response(
        user_command: str,
        excel_bytes: bytes,
        session_summary: Optional[str] = None
) -> ResponseResult:
    """
    LLM 서비스의 진입점 함수

    Args:
        user_command: 사용자가 입력한 자연어 명령
        excel_bytes: 현재 엑셀 파일의 바이트 데이터
        session_summary: 이전 대화 요약 (옵션)

    Returns:
        LLMResultInternal: LLM 응답 결과
    """
    service = LLMService()
    return service.get_llm_response(user_command, excel_bytes, session_summary)