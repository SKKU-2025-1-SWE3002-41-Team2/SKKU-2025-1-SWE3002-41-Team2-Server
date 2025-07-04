# app/services/excel.py
"""
엑셀 파일 조작 서비스
openpyxl을 사용하여 엑셀 파일을 직접 조작하는 기능을 제공합니다.
"""
import io
import re
from typing import List, Any, Optional, Union
from openpyxl import load_workbook, Workbook
from openpyxl.utils import column_index_from_string
import re

from app.schemas.excel_schema import ExcelCommand


class ExcelManipulator:
    """
    엑셀 파일을 조작하는 클래스
    명령어를 받아서 실제 엑셀 파일을 수정합니다.
    """

    def __init__(self):
        """ExcelManipulator 초기화"""
        self.workbook: Optional[Workbook] = None
        self.active_sheet = None

    def load_from_bytes(self, excel_bytes: bytes) -> None:
        """
        바이트 데이터에서 엑셀 파일을 로드합니다.

        Args:
            excel_bytes: 엑셀 파일의 바이트 데이터
        """
        self.workbook = load_workbook(io.BytesIO(excel_bytes))
        self.active_sheet = self.workbook.active

    def save_to_bytes(self) -> bytes:
        """
        현재 워크북을 바이트 데이터로 저장합니다.

        Returns:
            엑셀 파일의 바이트 데이터
        """
        if not self.workbook:
            raise ValueError("워크북이 로드되지 않았습니다.")

        output = io.BytesIO()
        self.workbook.save(output)
        output.seek(0)
        return output.getvalue()

    def execute_commands(self, commands: List[ExcelCommand]) -> None:
        """
        명령어 리스트를 순차적으로 실행합니다.

        Args:
            commands: 실행할 ExcelCommand 리스트
        """
        if not self.workbook or not self.active_sheet:
            raise ValueError("워크북이 로드되지 않았습니다.")

        for command in commands:
            self._execute_single_command(command)

    def _execute_single_command(self, command: ExcelCommand) -> None:
        """
        단일 명령어를 실행합니다.

        Args:
            command: 실행할 ExcelCommand
        """
        command_type = command.command_type.lower()

        # 함수 관련 명령어
        if command_type == "sum":
            self._apply_sum(command)
        elif command_type == "average":
            self._apply_average(command)
        elif command_type == "count":
            self._apply_count(command)
        elif command_type == "max":
            self._apply_max(command)
        elif command_type == "min":
            self._apply_min(command)

        # 텍스트 관련 명령어
        elif command_type == "left":
            self._apply_left(command)
        elif command_type == "right":
            self._apply_right(command)
        elif command_type == "mid":
            self._apply_mid(command)
        elif command_type == "len":
            self._apply_len(command)
        elif command_type == "round":
            self._apply_round(command)
        elif command_type == "isblank":
            self._apply_isblank(command)

        # 논리 관련 명령어
        elif command_type == "if":
            self._apply_if(command)
        elif command_type == "and":
            self._apply_logical_formula(command, "AND")
        elif command_type == "or":
            self._apply_logical_formula(command, "OR")

        # 검색 관련 명령어
        elif command_type == "vlookup":
            p = command.parameters
            formula = f'=VLOOKUP({p["lookup_value"]}, {p["table_array"]}, {p["col_index"]}, {str(p["range_lookup"]).upper()})'
            self.active_sheet[command.target_cell] = formula
        elif command_type == "hlookup":
            p = command.parameters
            formula = f'=HLOOKUP({p["lookup_value"]}, {p["table_array"]}, {p["row_index"]}, {str(p["range_lookup"]).upper()})'
            self.active_sheet[command.target_cell] = formula
        elif command_type == "index":
            p = command.parameters
            formula = f'=INDEX({p["array"]}, {p["row_num"]}, {p["col_num"]})'
            self.active_sheet[command.target_cell] = formula
        elif command_type == "match":
            p = command.parameters
            formula = f'=MATCH({p["lookup_value"]}, {p["lookup_array"]}, {p["match_type"]})'
            self.active_sheet[command.target_cell] = formula

        # 데이터 관련 명령어
        elif command_type == "set_value":
            self._set_value(command)
        elif command_type == "clear":
            self._clear_cells(command)
        elif command_type == "merge":
            self._merge_cells(command)
        elif command_type == "unmerge":
            self._unmerge_cells(command)

        # 조건부 함수 명령어
        elif command_type == "countif":
            self._apply_countif(command)
        elif command_type == "sumif":
            self._apply_sumif(command)
        elif command_type == "averageif":
            self._apply_averageif(command)

        # 텍스트 처리 함수 명령어
        elif command_type == "trim":
            self._apply_trim(command)
        elif command_type == "upper":
            self._apply_upper(command)
        elif command_type == "lower":
            self._apply_lower(command)
        elif command_type == "substitute":
            self._apply_substitute(command)

        # 고급 논리 함수
        elif command_type == "iferror":
            self._apply_iferror(command)
        elif command_type == "ifna":
            self._apply_ifna(command)
        elif command_type == "ifs":
            self._apply_ifs(command)

        # 고급 검색 함수
        elif command_type == "xlookup":
            self._apply_xlookup(command)
        elif command_type == "filter":
            self._apply_filter(command)
        elif command_type == "unique":
            self._apply_unique(command)

        # 통계 함수
        elif command_type == "median":
            self._apply_median(command)
        elif command_type == "mode":
            self._apply_mode(command)
        elif command_type == "stdev":
            self._apply_stdev(command)
        elif command_type == "rank":
            self._apply_rank(command)

        else:
            print(f"지원하지 않는 명령어: {command_type}")

    # ──────────────────────────────
    # 수식 함수
    # ──────────────────────────────
    def _apply_sum(self, command: ExcelCommand) -> None:
        """SUM 함수를 적용합니다."""
        if command.parameters and "range" in command.parameters:
            range_str = command.parameters["range"]
            formula = f"=SUM({range_str})"
            self.active_sheet[command.target_cell] = formula

    def _apply_average(self, command: ExcelCommand) -> None:
        """AVERAGE 함수를 적용합니다."""
        if command.parameters and "range" in command.parameters:
            range_str = command.parameters["range"]
            formula = f"=AVERAGE({range_str})"
            self.active_sheet[command.target_cell] = formula

    def _apply_count(self, command: ExcelCommand) -> None:
        """COUNT 함수를 적용합니다."""
        if command.parameters and "range" in command.parameters:
            range_str = command.parameters["range"]
            formula = f"=COUNT({range_str})"
            self.active_sheet[command.target_cell] = formula

    def _apply_max(self, command: ExcelCommand) -> None:
        """MAX 함수를 적용합니다."""
        if command.parameters and "range" in command.parameters:
            range_str = command.parameters["range"]
            formula = f"=MAX({range_str})"
            self.active_sheet[command.target_cell] = formula

    def _apply_min(self, command: ExcelCommand) -> None:
        """MIN 함수를 적용합니다."""
        if command.parameters and "range" in command.parameters:
            range_str = command.parameters["range"]
            formula = f"=MIN({range_str})"
            self.active_sheet[command.target_cell] = formula

    def _apply_concatenate(self, command: ExcelCommand):
        """CONCATENATE 함수를 적용합니다."""
        values = command.parameters.get("values", [])
        if not values:
            return
        # 각 값을 셀 참조나 문자열로 처리
        arg_str = ",".join(str(v) for v in values)
        self.active_sheet[command.target_cell] = f"=CONCATENATE({arg_str})"

    def _apply_left(self, command: ExcelCommand):
        """LEFT 함수를 적용합니다."""
        text = command.parameters.get("text", "")
        num_chars = command.parameters.get("num_chars", 1)
        if not text:
            return
        self.active_sheet[command.target_cell] = f"=LEFT({text},{num_chars})"

    # ──────────────────────────────
    # 조건부 함수
    # ──────────────────────────────
    def _apply_countif(self, command: ExcelCommand) -> None:
        """COUNTIF 함수를 적용합니다."""
        if command.parameters and "range" in command.parameters and "criteria" in command.parameters:
            range_str = command.parameters["range"]
            criteria = command.parameters["criteria"]
            formula = f"=COUNTIF({range_str}, {criteria})"
            self.active_sheet[command.target_cell] = formula

    def _apply_right(self, command: ExcelCommand):
        """RIGHT 함수를 적용합니다."""
        text = command.parameters.get("text", "")
        num_chars = command.parameters.get("num_chars", 1)
        if not text:
            return
        self.active_sheet[command.target_cell] = f"=RIGHT({text},{num_chars})"

    def _apply_sumif(self, command: ExcelCommand) -> None:
        """SUMIF 함수를 적용합니다."""
        if command.parameters and "range" in command.parameters and "criteria" in command.parameters:
            range_str = command.parameters["range"]
            criteria = command.parameters["criteria"]
            sum_range = command.parameters.get("sum_range", range_str)
            formula = f"=SUMIF({range_str}, {criteria}, {sum_range})"
            self.active_sheet[command.target_cell] = formula

    def _apply_averageif(self, command: ExcelCommand) -> None:
        """AVERAGEIF 함수를 적용합니다."""
        if command.parameters and "range" in command.parameters and "criteria" in command.parameters:
            range_str = command.parameters["range"]
            criteria = command.parameters["criteria"]
            avg_range = command.parameters.get("avg_range", range_str)
            formula = f"=AVERAGEIF({range_str}, {criteria}, {avg_range})"
            self.active_sheet[command.target_cell] = formula

    def _apply_mid(self, command: ExcelCommand):
        """MID 함수를 적용합니다."""
        text = command.parameters.get("text", "")
        start_num = command.parameters.get("start_num", 1)
        num_chars = command.parameters.get("num_chars", 1)
        if not text:
            return
        self.active_sheet[command.target_cell] = f"=MID({text},{start_num},{num_chars})"

    def _apply_len(self, command: ExcelCommand):
        """LEN 함수를 적용합니다."""
        text = command.parameters.get("text", "")
        if not text:
            return
        self.active_sheet[command.target_cell] = f"=LEN({text})"

    def _apply_round(self, command: ExcelCommand) -> None:
        """
        ROUND 함수를 적용합니다.
        기존 셀의 값을 그대로 사용하여 지정된 소수점 자리수로 반올림합니다.

        Args:
            command: ExcelCommand 객체
                - target_cell: 반올림을 적용할 셀 범위
                - parameters["num_digits"]: 반올림할 소수점 자릿수
        """
        # parameters가 딕셔너리이므로 키로 접근
        num_digits = command.parameters.get("num_digits", 0)  # 기본값 0

        def apply_round_to_cell(cell):
            # 현재 셀의 값 가져오기
            current_value = cell.value

            # 값이 없으면 스킵
            if current_value is None:
                return

            # 값을 문자열로 변환하여 처리
            value_str = str(current_value)

            # 수식인지 확인 (=로 시작하는 경우)
            if value_str.startswith('='):
                # = 기호 제거하고 수식 내용만 추출
                formula_content = str(value_str)
                # 앞의 =을 모두 제거
                while formula_content.startswith('='):
                    formula_content = formula_content[1:]
                # print(f"수식 감지: {formula_content} (소수점 자리수: {num_digits}) 기존 값: {value_str}")
                # ROUND 함수로 감싸기
                new_formula = f"=ROUND({formula_content}, {num_digits})"
            else:
                # 일반 값인 경우 그대로 사용
                new_formula = f"=ROUND({current_value}, {num_digits})"

            # 새로운 수식 적용
            cell.value = new_formula

        # 범위에 함수 적용
        self._apply_to_range(command.target_cell, apply_round_to_cell)

    def _apply_isblank(self, command: ExcelCommand):
        """ISBLANK 함수를 적용합니다."""
        value = command.parameters.get("value", "")
        if not value:
            return
        self.active_sheet[command.target_cell] = f"=ISBLANK({value})"

    def _apply_if(self, command: ExcelCommand) -> None:
        c = command.parameters
        formula = f'=IF({c["condition"]}, "{c["true_value"]}", "{c["false_value"]}")'
        self.active_sheet[command.target_cell] = formula

    def _apply_logical_formula(self, command: ExcelCommand, func_name: str) -> None:
        conditions = command.parameters.get("conditions", [])
        joined = ",".join(conditions)
        formula = f"={func_name.upper()}({joined})"
        self.active_sheet[command.target_cell] = formula

    # ──────────────────────────────
    # 텍스트 처리 함수
    # ──────────────────────────────
    def _apply_trim(self, command: ExcelCommand) -> None:
        """TRIM 함수를 적용합니다."""
        if command.parameters and "source" in command.parameters:
            source = command.parameters["source"]
            formula = f"=TRIM({source})"
            self.active_sheet[command.target_cell] = formula

    # 데이터 관련 명령어 구현
    def _set_value(self, command: ExcelCommand) -> None:
        """셀에 값을 설정합니다."""
        if command.parameters and "value" in command.parameters:
            value = command.parameters["value"]

            if ":" in command.target_cell:
                # 범위의 모든 셀에 같은 값 설정
                self._apply_to_range(command.target_cell, lambda cell: setattr(cell, 'value', value))
            else:
                # 단일 셀에 값 설정
                self.active_sheet[command.target_cell] = value

    def _clear_cells(self, command: ExcelCommand) -> None:
        """셀의 내용을 지웁니다."""
        self._apply_to_range(command.target_cell, lambda cell: setattr(cell, 'value', None))

    def _merge_cells(self, command: ExcelCommand) -> None:
        """셀을 병합합니다."""
        self.active_sheet.merge_cells(command.target_cell)

    def _unmerge_cells(self, command: ExcelCommand) -> None:
        """셀 병합을 해제합니다."""
        self.active_sheet.unmerge_cells(command.target_cell)


    def _apply_to_range(self, target_cell: str, func) -> None:
        """범위의 모든 셀에 함수를 적용하는 헬퍼 메서드"""
        if ":" in target_cell:
            # 범위인 경우
            for row in self.active_sheet[target_cell]:
                for cell in row:
                    func(cell)
        else:
            # 단일 셀인 경우
            cell = self.active_sheet[target_cell]
            func(cell)

    def _parse_range(self, range_str: str) -> tuple:
        """
        셀 범위 문자열을 파싱합니다.

        Args:
            range_str: 셀 범위 (예: "A1:B10")

        Returns:
            (start_col, start_row, end_col, end_row) 튜플
        """
        pattern = r'([A-Z]+)(\d+)'

        if ":" in range_str:
            start, end = range_str.split(":")
            start_match = re.match(pattern, start)
            end_match = re.match(pattern, end)

            if start_match and end_match:
                start_col = column_index_from_string(start_match.group(1))
                start_row = int(start_match.group(2))
                end_col = column_index_from_string(end_match.group(1))
                end_row = int(end_match.group(2))
                return (start_col, start_row, end_col, end_row)
        else:
            match = re.match(pattern, range_str)
            if match:
                col = column_index_from_string(match.group(1))
                row = int(match.group(2))
                return (col, row, col, row)

        raise ValueError(f"잘못된 셀 범위 형식: {range_str}")

    def _apply_iferror(self, command: ExcelCommand) -> None:
        """
        IFERROR 함수를 적용합니다.
        오류가 발생하면 지정된 값을 반환합니다.

        Args:
            command: ExcelCommand 객체
                - parameters[0]: 검사할 수식 또는 범위
                - parameters[1]: 오류 시 반환할 값
                "test_formula": parameters[0],
                "error_value": parameters[1]
        """
        if command.parameters and len(command.parameters) >= 2:
            test_formula = command.parameters["test_formula"]
            error_value = command.parameters["error_value"]

            # IFERROR 수식 생성
            formula = f"=IFERROR({test_formula}, {error_value})"
            self.active_sheet[command.target_cell] = formula

    def _apply_ifna(self, command: ExcelCommand) -> None:
        """
        IFNA 함수를 적용합니다.
        #N/A 오류가 발생하면 지정된 값을 반환합니다.

        Args:
            command: ExcelCommand 객체
                - parameters[0]: 검사할 수식 또는 범위
                - parameters[1]: #N/A 오류 시 반환할 값
        """
        if command.parameters and len(command.parameters) >= 2:
            test_formula = command.parameters["test_formula"]
            na_value = command.parameters["na_value"]

            # IFNA 수식 생성
            formula = f"=IFNA({test_formula}, {na_value})"
            self.active_sheet[command.target_cell] = formula

    def _apply_ifs(self, command: ExcelCommand) -> None:
        """
        IFS 함수를 적용합니다.
        여러 조건을 순차적으로 검사하여 첫 번째 참인 조건의 결과를 반환합니다.
        조건과 값의 데이터 타입을 자동으로 판별하여 올바른 Excel 수식을 생성합니다.

        Args:
            command: ExcelCommand 객체
                - target_cell: IFS 함수를 적용할 셀
                - parameters["conditions_values"]: [조건1, 값1, 조건2, 값2, ...] 형태의 배열

        Example:
            조건-값 쌍: ["B2>=90", "A", "B2>=80", "B", "TRUE", "F"]
            생성되는 수식: =IFS(B2>=90,"A",B2>=80,"B",TRUE,"F")
        """
        if not command.parameters or "conditions_values" not in command.parameters:
            print(f"[IFS 오류] conditions_values 파라미터가 없습니다. target_cell: {command.target_cell}")
            return

        conditions_values = command.parameters["conditions_values"]

        # 조건-값 쌍 검증
        if len(conditions_values) < 2:
            print(f"[IFS 오류] 최소 하나의 조건-값 쌍이 필요합니다. target_cell: {command.target_cell}")
            return

        if len(conditions_values) % 2 != 0:
            print(f"[IFS 오류] 조건과 값이 쌍으로 제공되어야 합니다. 현재 개수: {len(conditions_values)}")
            return

        # 조건-값 쌍들을 처리하여 Excel 수식 생성
        formula_parts = []

        try:
            for i in range(0, len(conditions_values), 2):
                condition = conditions_values[i]
                value = conditions_values[i + 1]

                # 조건 부분 처리 (일반적으로 수식이므로 그대로 사용)
                processed_condition = self._process_ifs_condition(condition)

                # 값 부분 처리 (타입에 따라 다르게 처리)
                processed_value = self._process_ifs_value(value)

                formula_parts.append(f"{processed_condition},{processed_value}")

                # 디버깅 로그
                # print(f"[IFS 처리] 조건 {i // 2 + 1}: {condition} -> {processed_condition}")
                # print(f"[IFS 처리] 값 {i // 2 + 1}: {value} ({type(value).__name__}) -> {processed_value}")

            # 최종 IFS 수식 생성 및 적용
            if formula_parts:
                formula = f"=IFS({','.join(formula_parts)})"
                # print(f"[IFS 완료] 생성된 수식: {formula}")
                self.active_sheet[command.target_cell] = formula

        except Exception as e:
            print(f"[IFS 오류] 수식 생성 중 오류 발생: {str(e)}")
            print(f"[IFS 오류] target_cell: {command.target_cell}, parameters: {command.parameters}")

    def _process_ifs_condition(self, condition: Union[str, int, float, bool]) -> str:
        """
        IFS 함수의 조건 부분을 처리합니다.
        조건은 대부분 논리식이므로 문자열로 변환하여 반환합니다.

        Args:
            condition: 조건 (예: "A2>=90", "AND(A2>80,B2>0)", "TRUE")

        Returns:
            처리된 조건 문자열
        """
        if condition is None:
            return "FALSE"

        # 불린값 처리
        if isinstance(condition, bool):
            return str(condition).upper()

        # 문자열이 아닌 경우 문자열로 변환
        return str(condition)

    def _process_ifs_value(self, value: Union[str, int, float, bool]) -> str:
        """
        IFS 함수의 값 부분을 처리합니다.
        데이터 타입에 따라 Excel에서 올바르게 인식할 수 있는 형태로 변환합니다.

        Args:
            value: 값 (문자열, 숫자, 수식 등)

        Returns:
            Excel에서 사용할 수 있는 형태로 처리된 값
        """
        # 1. None 값 처리
        if value is None:
            return '""'

        # 2. 불린값 처리
        if isinstance(value, bool):
            return str(value).upper()  # TRUE 또는 FALSE

        # 3. 숫자 처리 (정수, 실수)
        if isinstance(value, (int, float)):
            return str(value)

        # 4. 문자열 처리 - 가장 복잡한 부분
        if isinstance(value, str):
            return self._process_ifs_string_value(value)

        # 5. 기타 타입은 문자열로 변환하여 따옴표로 감싸기
        return f'"{str(value)}"'

    def _process_ifs_string_value(self, value: str) -> str:
        """
        IFS 함수의 문자열 값을 분석하여 적절한 형태로 처리합니다.

        Args:
            value: 문자열 값

        Returns:
            처리된 값
        """
        # 빈 문자열 처리
        if not value.strip():
            return '""'

        # 이미 따옴표로 감싸진 문자열 (중복 처리 방지)
        if value.startswith('"') and value.endswith('"') and len(value) > 1:
            return value

        # 1. TRUE/FALSE 문자열인지 확인 (대소문자 무관)
        if value.upper() in ['TRUE', 'FALSE']:
            return value.upper()

        # 2. Excel 함수인지 확인
        if self._is_excel_function(value):
            return value

        # 3. 셀 참조인지 확인 (예: A1, B2, A1:B10, Sheet1!A1)
        if self._is_cell_reference(value):
            return value

        # 4. 수식인지 확인 (=로 시작하거나 연산자 포함)
        if self._is_formula_expression(value):
            return value

        # 5. 숫자 문자열인지 확인
        if self._is_numeric_string(value):
            return value  # 숫자는 따옴표 없이

        # 6. 일반 텍스트는 따옴표로 감싸기
        return f'"{value}"'

    def _is_excel_function(self, value: str) -> bool:
        """
        Excel 함수인지 판별합니다.

        Args:
            value: 검사할 문자열

        Returns:
            Excel 함수이면 True
        """
        # Excel 함수 목록 (주요 함수들)
        excel_functions = {
            # 수학 함수
            'SUM', 'AVERAGE', 'COUNT', 'COUNTA', 'MAX', 'MIN', 'ROUND', 'ROUNDUP', 'ROUNDDOWN',
            'ABS', 'SQRT', 'POWER', 'MOD', 'INT', 'CEILING', 'FLOOR',

            # 논리 함수
            'IF', 'AND', 'OR', 'NOT', 'IFS', 'IFERROR', 'IFNA', 'IFBLANK',

            # 텍스트 함수
            'CONCATENATE', 'LEFT', 'RIGHT', 'MID', 'LEN', 'TRIM', 'UPPER', 'LOWER',
            'SUBSTITUTE', 'REPLACE', 'FIND', 'SEARCH', 'EXACT',

            # 날짜/시간 함수
            'TODAY', 'NOW', 'YEAR', 'MONTH', 'DAY', 'DATE', 'TIME', 'HOUR', 'MINUTE', 'SECOND',

            # 검색/참조 함수
            'VLOOKUP', 'HLOOKUP', 'INDEX', 'MATCH', 'LOOKUP', 'CHOOSE', 'XLOOKUP', 'FILTER', 'UNIQUE',

            # 정보 함수
            'ISBLANK', 'ISNUMBER', 'ISTEXT', 'ISERROR', 'ISNA', 'ISODD', 'ISEVEN',

            # 통계 함수
            'MEDIAN', 'MODE', 'STDEV', 'VAR', 'RANK', 'PERCENTILE', 'QUARTILE',

            # 조건부 함수
            'COUNTIF', 'COUNTIFS', 'SUMIF', 'SUMIFS', 'AVERAGEIF', 'AVERAGEIFS'
        }

        # 함수명(매개변수) 패턴 검사
        # 예: SUM(A1:A10), CONCATENATE(A1," ",B1) 등
        function_pattern = r'^([A-Z_]+)\s*\('
        match = re.match(function_pattern, value.upper())

        if match:
            function_name = match.group(1)
            return function_name in excel_functions

        return False

    def _is_cell_reference(self, value: str) -> bool:
        """
        셀 참조인지 판별합니다.

        Args:
            value: 검사할 문자열

        Returns:
            셀 참조이면 True
        """
        try:
            # 1. 단일 셀 참조: A1, B2, Z99, AA1, AB123 등
            single_cell_pattern = r'^[A-Z]{1,3}\d{1,7}$'

            # 2. 범위 참조: A1:B10, C2:D5, AA1:AB100 등
            range_pattern = r'^[A-Z]{1,3}\d{1,7}:[A-Z]{1,3}\d{1,7}$'

            # 3. 시트 참조: Sheet1!A1, 'Sheet Name'!A1:B10 등
            sheet_reference_pattern = r'^[\'"]?[\w\s]+[\'"]?![A-Z]{1,3}\d{1,7}(:[A-Z]{1,3}\d{1,7})?$'

            # 4. 전체 열/행 참조: A:A, B:D, 1:1, 1:10 등
            full_column_pattern = r'^[A-Z]{1,3}:[A-Z]{1,3}$'
            full_row_pattern = r'^\d+:\d+$'

            # 모든 패턴 검사
            patterns = [
                single_cell_pattern,
                range_pattern,
                sheet_reference_pattern,
                full_column_pattern,
                full_row_pattern
            ]

            value_upper = value.upper()
            for pattern in patterns:
                if re.match(pattern, value_upper):
                    return True

            return False

        except Exception:
            return False

    def _is_formula_expression(self, value: str) -> bool:
        """
        수식 표현인지 판별합니다.

        Args:
            value: 검사할 문자열

        Returns:
            수식이면 True
        """
        # =로 시작하는 경우
        if value.startswith('='):
            return True

        # 산술 연산자가 포함되고 셀 참조도 포함된 경우
        # 예: A2*0.1, B2+C2, (A1+B1)/2 등
        arithmetic_operators = ['+', '-', '*', '/', '^', '%']
        has_operator = any(op in value for op in arithmetic_operators)

        # 셀 참조 패턴 확인
        has_cell_ref = bool(re.search(r'[A-Z]{1,3}\d{1,7}', value.upper()))

        # 문자열 연결 연산자 & 확인
        has_concat = '&' in value

        return (has_operator or has_concat) and has_cell_ref

    def _is_numeric_string(self, value: str) -> bool:
        """
        숫자 문자열인지 판별합니다.

        Args:
            value: 검사할 문자열

        Returns:
            숫자 문자열이면 True
        """
        try:
            # 정수 또는 실수로 변환 가능한지 확인
            float(value)
            return True
        except ValueError:
            return False

    def _apply_xlookup(self, command: ExcelCommand) -> None:
        """
        XLOOKUP 함수를 적용합니다.
        VLOOKUP의 개선된 버전으로 더 유연한 검색이 가능합니다.

        Args:
            command: ExcelCommand 객체
                - parameters["lookup_value"]: 찾을 값
                - parameters["lookup_array"]: 찾을 범위
                - parameters["return_array"]: 반환할 범위
                - parameters["if_not_found"]: (선택) 찾지 못했을 때 반환할 값
                - parameters["match_mode"]: (선택) 일치 모드
                - parameters["search_mode"]: (선택) 검색 모드
        """
        if command.parameters and "lookup_value" in command.parameters and "lookup_array" in command.parameters and "return_array" in command.parameters:
            lookup_value = command.parameters["lookup_value"]
            lookup_array = command.parameters["lookup_array"]
            return_array = command.parameters["return_array"]

            # 기본 XLOOKUP 수식
            formula_parts = [lookup_value, lookup_array, return_array]

            # 선택적 매개변수 추가
            if "if_not_found" in command.parameters:
                formula_parts.append(command.parameters["if_not_found"])
            if "match_mode" in command.parameters:
                formula_parts.append(command.parameters["match_mode"])
            if "search_mode" in command.parameters:
                formula_parts.append(command.parameters["search_mode"])

            formula = f"=XLOOKUP({', '.join(map(str, formula_parts))})"
            self.active_sheet[command.target_cell] = formula


    def _apply_filter(self, command: ExcelCommand) -> None:
        """
        FILTER 함수를 적용합니다.
        조건에 맞는 데이터만 필터링하여 반환합니다.

        Args:
            command: ExcelCommand 객체
                - parameters["array"]: 필터링할 범위
                - parameters["include"]: 조건
                - parameters["if_empty"]: (선택) 조건에 맞는 값이 없을 때 반환할 값
        """
        if command.parameters and "array" in command.parameters and "include" in command.parameters:
            array = command.parameters["array"]
            include = command.parameters["include"]

            if "if_empty" in command.parameters:
                if_empty = command.parameters["if_empty"]
                formula = f"=FILTER({array}, {include}, {if_empty})"
            else:
                formula = f"=FILTER({array}, {include})"

            self.active_sheet[command.target_cell] = formula

    def _apply_unique(self, command: ExcelCommand) -> None:
        """
        UNIQUE 함수를 적용합니다.
        중복을 제거한 고유값만 반환합니다.

        Args:
            command: ExcelCommand 객체
                - parameters["array"]: 고유값을 추출할 범위
                - parameters["by_col"]: (선택) True면 열 기준, False면 행 기준
                - parameters["exactly_once"]: (선택) True면 정확히 한 번만 나타나는 값만 반환
        """
        if command.parameters and "array" in command.parameters:
            array = command.parameters["array"]

            # 기본 UNIQUE 수식
            formula_parts = [array]

            # 선택적 매개변수 추가
            if "by_col" in command.parameters:
                by_col = command.parameters["by_col"]
                formula_parts.append(str(by_col).upper())

                if "exactly_once" in command.parameters:
                    exactly_once = command.parameters["exactly_once"]
                    formula_parts.append(str(exactly_once).upper())

            formula = f"=UNIQUE({', '.join(map(str, formula_parts))})"
            self.active_sheet[command.target_cell] = formula

    # 통계 함수 관련 메소드들
    def _apply_median(self, command: ExcelCommand) -> None:
        """
        MEDIAN 함수를 적용합니다.
        중간값(중위수)을 계산합니다.

        Args:
            command: ExcelCommand 객체
                - parameters[0]: 중위수를 계산할 범위
        """
        if command.parameters and "range" in command.parameters:
            range_str = command.parameters["range"]
            formula = f"=MEDIAN({range_str})"
            self.active_sheet[command.target_cell] = formula

    def _apply_mode(self, command: ExcelCommand) -> None:
        """
        MODE 함수를 적용합니다.
        가장 자주 나타나는 값(최빈값)을 반환합니다.
        참고: Excel 2010 이후 MODE.SNGL 사용 권장

        Args:
            command: ExcelCommand 객체
                - parameters[0]: 최빈값을 계산할 범위
        """
        if command.parameters and "range" in command.parameters:
            range_str = command.parameters["range"]
            # MODE.SNGL 사용 (Excel 2010 이후 권장)
            formula = f"=MODE.SNGL({range_str})"
            self.active_sheet[command.target_cell] = formula

    def _apply_stdev(self, command: ExcelCommand) -> None:
        """
        STDEV 함수를 적용합니다.
        표준편차를 계산합니다.
        참고: STDEV.S는 표본 표준편차, STDEV.P는 모집단 표준편차

        Args:
            command: ExcelCommand 객체
                - parameters[0]: 표준편차를 계산할 범위
                - parameters[1]: (선택) "S" 또는 "P" - 표본/모집단 구분
        """
        if command.parameters and "range" in command.parameters:
            range_str = command.parameters["range"]

            # 표본/모집단 구분 (기본값: 표본)
            stdev_type = command.parameters.get("type", "S")

            if stdev_type.upper() == "P":
                formula = f"=STDEV.P({range_str})"
            else:
                formula = f"=STDEV.S({range_str})"

            self.active_sheet[command.target_cell] = formula

    def _apply_rank(self, command: ExcelCommand) -> None:
        """
        RANK 함수를 적용합니다.
        특정 값이 전체에서 몇 번째 순위인지 반환합니다.
        참고: RANK.EQ 사용 (Excel 2010 이후 권장)

        Args:
            command: ExcelCommand 객체
                - parameters[0]: 순위를 구할 값 또는 셀 참조
                - parameters[1]: 비교할 범위
                - parameters[2]: (선택) 순서 - 0 또는 생략: 내림차순, 1: 오름차순
        """
        if command.parameters and len(command.parameters) >= 2:
            number = command.parameters["number"]
            ref = command.parameters["ref"]

            # 순서 매개변수 (기본값: 0 - 내림차순)
            order = command.parameters["order"] if len(command.parameters) > 2 else "0"

            formula = f"=RANK.EQ({number}, {ref}, {order})"
            self.active_sheet[command.target_cell] = formula

    def log_worksheet_contents(self, log_title: str = "워크시트 내용") -> None:
        """
        현재 워크시트의 모든 내용을 로그에 출력합니다.

        Args:
            log_title: 로그 제목
        """
        if not self.workbook or not self.active_sheet:
            print(f"[{log_title}] 워크북이 로드되지 않았습니다.")
            return

        # print(f"\n{'=' * 50}")
        # print(f"[{log_title}]")
        # print(f"{'=' * 50}")

        # 워크시트 기본 정보
        ws = self.active_sheet
        max_row = ws.max_row
        max_col = ws.max_column

        # print(f"워크시트명: {ws.title}")
        # print(f"최대 행: {max_row}, 최대 열: {max_col}")

        if max_row == 1 and max_col == 1 and ws.cell(1, 1).value is None:
            print("워크시트가 비어있습니다.")
            print(f"{'=' * 50}\n")
            return

        '''
        # 헤더 출력 (열 번호)
        print("\n   ", end="")
        for col in range(1, max_col + 1):
            col_letter = ws.cell(1, col).column_letter
            print(f"{col_letter:>12}", end="")
        print()
        
        # 각 행의 데이터 출력
        for row in range(1, max_row + 1):
            print(f"{row:>3}:", end="")

            for col in range(1, max_col + 1):
                cell = ws.cell(row, col)

                # 셀 값 가져오기
                value = cell.value

                # 값이 None인 경우 공백 표시
                if value is None:
                    display_value = ""
                # 수식인 경우 수식과 결과값 모두 표시
                elif isinstance(value, str) and value.startswith('='):
                    try:
                        # 수식의 계산 결과 시도
                        calculated_value = cell.displayed_value if hasattr(cell, 'displayed_value') else "계산필요"
                        display_value = f"{value}({calculated_value})"
                    except:
                        display_value = value
                else:
                    display_value = str(value)

                # 너무 긴 값은 잘라내기
                if len(display_value) > 10:
                    display_value = display_value[:7] + "..."

                print(f"{display_value:>12}", end="")
            print()  # 행 끝에서 줄바꿈

        print(f"{'=' * 50}\n")
        '''


def process_excel_with_commands(
        excel_bytes: bytes,
        commands: Any
) -> bytes:
    """
    엑셀 파일에 명령어를 적용하고 결과를 반환합니다.

    Args:
        excel_bytes: 원본 엑셀 파일의 바이트 데이터
        commands: 적용할 명령어 리스트

    Returns:
        수정된 엑셀 파일의 바이트 데이터
    """
    manipulator = ExcelManipulator()

    # 엑셀 파일 로드
    manipulator.load_from_bytes(excel_bytes)

    # 🔹 수정 전 상태 로그 출력
    manipulator.log_worksheet_contents("명령어 적용 전 워크시트 상태")

    # 명령어 실행
    print(f"\n[실행할 명령어 목록]")
    for i, command in enumerate(commands, 1):
        print(f"  {i}. {command.command_type} -> {command.target_cell} | {command.parameters}")
    print()

    manipulator.execute_commands(commands)

    manipulator.log_worksheet_contents("명령어 적용 후 워크시트 상태")

    # 결과 저장 및 반환
    return manipulator.save_to_bytes()


def create_empty_excel() -> bytes:
    """
    빈 엑셀 파일을 생성합니다.

    Returns:
        빈 엑셀 파일의 바이트 데이터
    """
    workbook = Workbook()
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()