# app/services/excel.py
"""
엑셀 파일 조작 서비스
openpyxl을 사용하여 엑셀 파일을 직접 조작하는 기능을 제공합니다.
"""
import io
from typing import List, Any, Optional
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import column_index_from_string
import re

from app.schemas.excel import ExcelCommand


class ExcelManipulator:
    """
    엑셀 파일을 조작하는 클래스
    명령어를 받아서 실제 엑셀 파일을 수정합니다.
    """

    def __init__(self):
        """ExcelManipulator 초기화"""
        self.workbook: Optional[Workbook] = None
        self.active_sheet = None

    def load_from_bytes(self, excel_bytes: bytes) -> None:
        """
        바이트 데이터에서 엑셀 파일을 로드합니다.

        Args:
            excel_bytes: 엑셀 파일의 바이트 데이터
        """
        self.workbook = load_workbook(io.BytesIO(excel_bytes))
        self.active_sheet = self.workbook.active

    def save_to_bytes(self) -> bytes:
        """
        현재 워크북을 바이트 데이터로 저장합니다.

        Returns:
            엑셀 파일의 바이트 데이터
        """
        if not self.workbook:
            raise ValueError("워크북이 로드되지 않았습니다.")

        output = io.BytesIO()
        self.workbook.save(output)
        output.seek(0)
        return output.getvalue()

    def execute_commands(self, commands: List[ExcelCommand]) -> None:
        """
        명령어 리스트를 순차적으로 실행합니다.

        Args:
            commands: 실행할 ExcelCommand 리스트
        """
        if not self.workbook or not self.active_sheet:
            raise ValueError("워크북이 로드되지 않았습니다.")

        for command in commands:
            self._execute_single_command(command)

    def _execute_single_command(self, command: ExcelCommand) -> None:
        """
        단일 명령어를 실행합니다.

        Args:
            command: 실행할 ExcelCommand
        """
        command_type = command.command_type.lower()

        # 함수 관련 명령어
        if command_type == "sum":
            self._apply_sum(command)
        elif command_type == "average":
            self._apply_average(command)
        elif command_type == "count":
            self._apply_count(command)
        elif command_type == "max":
            self._apply_max(command)
        elif command_type == "min":
            self._apply_min(command)


        # 데이터 관련 명령어
        elif command_type == "set_value":
            self._set_value(command)
        elif command_type == "clear":
            self._clear_cells(command)
        elif command_type == "merge":
            self._merge_cells(command)
        elif command_type == "unmerge":
            self._unmerge_cells(command)

        # ----- 조건부 함수 -----
        elif command_type == "countif":
            self._apply_countif(command)
        elif command_type == "sumif":
            self._apply_sumif(command)
        elif command_type == "averageif":
            self._apply_averageif(command)

        # ----- 텍스트 처리 함수 -----
        elif command_type == "trim":
            self._apply_trim(command)
        elif command_type == "upper":
            self._apply_upper(command)
        elif command_type == "lower":
            self._apply_lower(command)
        elif command_type == "substitute":
            self._apply_substitute(command)
        # 고급 논리 함수
        elif command_type == "iferror":
            self._apply_iferror(command)
        elif command_type == "ifna":
            self._apply_ifna(command)
        elif command_type == "ifs":
            self._apply_ifs(command)

        # 고급 검색 함수
        elif command_type == "xlookup":
            self._apply_xlookup(command)
        elif command_type == "filter":
            self._apply_filter(command)
        elif command_type == "unique":
            self._apply_unique(command)

        # 통계 함수
        elif command_type == "median":
            self._apply_median(command)
        elif command_type == "mode":
            self._apply_mode(command)
        elif command_type == "stdev":
            self._apply_stdev(command)
        elif command_type == "rank":
            self._apply_rank(command)

        else:
            print(f"지원하지 않는 명령어: {command_type}")

    # 함수 관련 명령어 구현
    def _apply_sum(self, command: ExcelCommand) -> None:
        """SUM 함수를 적용합니다."""
        if command.parameters and "range" in command.parameters:
            range_str = command.parameters["range"]
            formula = f"=SUM({range_str})"
            self.active_sheet[command.target_range] = formula

    def _apply_average(self, command: ExcelCommand) -> None:
        """AVERAGE 함수를 적용합니다."""
        if command.parameters and "range" in command.parameters:
            range_str = command.parameters["range"]
            formula = f"=AVERAGE({range_str})"
            self.active_sheet[command.target_range] = formula

    def _apply_count(self, command: ExcelCommand) -> None:
        """COUNT 함수를 적용합니다."""
        if command.parameters and "range" in command.parameters:
            range_str = command.parameters["range"]
            formula = f"=COUNT({range_str})"
            self.active_sheet[command.target_range] = formula

    def _apply_max(self, command: ExcelCommand) -> None:
        """MAX 함수를 적용합니다."""
        if command.parameters and "range" in command.parameters:
            range_str = command.parameters["range"]
            formula = f"=MAX({range_str})"
            self.active_sheet[command.target_range] = formula

    def _apply_min(self, command: ExcelCommand) -> None:
        """MIN 함수를 적용합니다."""
        if command.parameters and "range" in command.parameters:
            range_str = command.parameters["range"]
            formula = f"=MIN({range_str})"
            self.active_sheet[command.target_range] = formula

    # ──────────────────────────────
    # 조건부 함수
    # ──────────────────────────────
    def _apply_countif(self, command: ExcelCommand) -> None:
        """COUNTIF 함수를 적용합니다."""
        if command.parameters and "range" in command.parameters and "criteria" in command.parameters:
            range_str = command.parameters["range"]
            criteria = command.parameters["criteria"]
            formula = f"=COUNTIF({range_str}, {criteria})"
            self.active_sheet[command.target_range] = formula

    def _apply_sumif(self, command: ExcelCommand) -> None:
        """SUMIF 함수를 적용합니다."""
        if command.parameters and "range" in command.parameters and "criteria" in command.parameters:
            range_str = command.parameters["range"]
            criteria = command.parameters["criteria"]
            sum_range = command.parameters.get("sum_range", range_str)
            formula = f"=SUMIF({range_str}, {criteria}, {sum_range})"
            self.active_sheet[command.target_range] = formula

    def _apply_averageif(self, command: ExcelCommand) -> None:
        """AVERAGEIF 함수를 적용합니다."""
        if command.parameters and "range" in command.parameters and "criteria" in command.parameters:
            range_str = command.parameters["range"]
            criteria = command.parameters["criteria"]
            avg_range = command.parameters.get("avg_range", range_str)
            formula = f"=AVERAGEIF({range_str}, {criteria}, {avg_range})"
            self.active_sheet[command.target_range] = formula

    # ──────────────────────────────
    # 텍스트 처리 함수
    # ──────────────────────────────
    def _apply_trim(self, command: ExcelCommand) -> None:
        """TRIM 함수를 적용합니다."""
        if command.parameters and "source" in command.parameters:
            source = command.parameters["source"]
            formula = f"=TRIM({source})"
            self.active_sheet[command.target_range] = formula

    def _apply_upper(self, command: ExcelCommand) -> None:
        """UPPER 함수를 적용합니다."""
        if command.parameters and "source" in command.parameters:
            source = command.parameters["source"]
            formula = f"=UPPER({source})"
            self.active_sheet[command.target_range] = formula

    def _apply_lower(self, command: ExcelCommand) -> None:
        """LOWER 함수를 적용합니다."""
        if command.parameters and "source" in command.parameters:
            source = command.parameters["source"]
            formula = f"=LOWER({source})"
            self.active_sheet[command.target_range] = formula

    def _apply_substitute(self, command: ExcelCommand) -> None:
        """SUBSTITUTE 함수를 적용합니다."""
        if command.parameters and {"source", "old_text", "new_text", "instance_number"} <= command.parameters.keys():
            source = command.parameters["source"]
            old_text = command.parameters["old_text"]
            new_text = command.parameters["new_text"]
            instance_number = command.parameters["instance_number"]
            formula = f"=SUBSTITUTE({source}, {old_text}, {new_text}, {instance_number})"
            self.active_sheet[command.target_range] = formula


    # 데이터 관련 명령어 구현
    def _set_value(self, command: ExcelCommand) -> None:
        """셀에 값을 설정합니다."""
        if command.parameters and "value" in command.parameters:
            value = command.parameters["value"]

            if ":" in command.target_range:
                # 범위의 모든 셀에 같은 값 설정
                self._apply_to_range(command.target_range, lambda cell: setattr(cell, 'value', value))
            else:
                # 단일 셀에 값 설정
                self.active_sheet[command.target_range] = value

    def _clear_cells(self, command: ExcelCommand) -> None:
        """셀의 내용을 지웁니다."""
        self._apply_to_range(command.target_range, lambda cell: setattr(cell, 'value', None))

    def _merge_cells(self, command: ExcelCommand) -> None:
        """셀을 병합합니다."""
        self.active_sheet.merge_cells(command.target_range)

    def _unmerge_cells(self, command: ExcelCommand) -> None:
        """셀 병합을 해제합니다."""
        self.active_sheet.unmerge_cells(command.target_range)


    def _apply_to_range(self, target_range: str, func) -> None:
        """범위의 모든 셀에 함수를 적용하는 헬퍼 메서드"""
        if ":" in target_range:
            # 범위인 경우
            for row in self.active_sheet[target_range]:
                for cell in row:
                    func(cell)
        else:
            # 단일 셀인 경우
            cell = self.active_sheet[target_range]
            func(cell)

    def _parse_range(self, range_str: str) -> tuple:
        """
        셀 범위 문자열을 파싱합니다.

        Args:
            range_str: 셀 범위 (예: "A1:B10")

        Returns:
            (start_col, start_row, end_col, end_row) 튜플
        """
        pattern = r'([A-Z]+)(\d+)'

        if ":" in range_str:
            start, end = range_str.split(":")
            start_match = re.match(pattern, start)
            end_match = re.match(pattern, end)

            if start_match and end_match:
                start_col = column_index_from_string(start_match.group(1))
                start_row = int(start_match.group(2))
                end_col = column_index_from_string(end_match.group(1))
                end_row = int(end_match.group(2))
                return (start_col, start_row, end_col, end_row)
        else:
            match = re.match(pattern, range_str)
            if match:
                col = column_index_from_string(match.group(1))
                row = int(match.group(2))
                return (col, row, col, row)

        raise ValueError(f"잘못된 셀 범위 형식: {range_str}")

    def _apply_iferror(self, command: ExcelCommand) -> None:
        """
        IFERROR 함수를 적용합니다.
        오류가 발생하면 지정된 값을 반환합니다.

        Args:
            command: ExcelCommand 객체
                - parameters[0]: 검사할 수식 또는 범위
                - parameters[1]: 오류 시 반환할 값
        """
        if command.parameters and len(command.parameters) >= 2:
            test_formula = command.parameters[0]
            error_value = command.parameters[1]

            # IFERROR 수식 생성
            formula = f"=IFERROR({test_formula}, {error_value})"
            self.active_sheet[command.target_range] = formula

    def _apply_ifna(self, command: ExcelCommand) -> None:
        """
        IFNA 함수를 적용합니다.
        #N/A 오류가 발생하면 지정된 값을 반환합니다.

        Args:
            command: ExcelCommand 객체
                - parameters[0]: 검사할 수식 또는 범위
                - parameters[1]: #N/A 오류 시 반환할 값
        """
        if command.parameters and len(command.parameters) >= 2:
            test_formula = command.parameters[0]
            na_value = command.parameters[1]

            # IFNA 수식 생성
            formula = f"=IFNA({test_formula}, {na_value})"
            self.active_sheet[command.target_range] = formula

    def _apply_ifs(self, command: ExcelCommand) -> None:
        """
        IFS 함수를 적용합니다.
        여러 조건을 순차적으로 검사하여 첫 번째 참인 조건의 결과를 반환합니다.

        Args:
            command: ExcelCommand 객체
                - parameters: [조건1, 값1, 조건2, 값2, ...] 형태의 배열
        """
        if command.parameters and len(command.parameters) >= 2:
            # 조건과 값의 쌍으로 수식 구성
            conditions_values = []
            for i in range(0, len(command.parameters), 2):
                if i + 1 < len(command.parameters):
                    condition = command.parameters[i]
                    value = command.parameters[i + 1]
                    conditions_values.append(f"{condition}, {value}")

            if conditions_values:
                formula = f"=IFS({', '.join(conditions_values)})"
                self.active_sheet[command.target_range] = formula

    # 고급 검색 함수 관련 메소드들
    def _apply_xlookup(self, command: ExcelCommand) -> None:
        """
        XLOOKUP 함수를 적용합니다.
        VLOOKUP의 개선된 버전으로 더 유연한 검색이 가능합니다.

        Args:
            command: ExcelCommand 객체
                - parameters[0]: 찾을 값
                - parameters[1]: 찾을 범위
                - parameters[2]: 반환할 범위
                - parameters[3]: (선택) 찾지 못했을 때 반환할 값
                - parameters[4]: (선택) 일치 모드
                - parameters[5]: (선택) 검색 모드
        """
        if command.parameters and len(command.parameters) >= 3:
            lookup_value = command.parameters[0]
            lookup_array = command.parameters[1]
            return_array = command.parameters[2]

            # 기본 XLOOKUP 수식
            formula_parts = [lookup_value, lookup_array, return_array]

            # 선택적 매개변수 추가
            if len(command.parameters) > 3:
                formula_parts.extend(command.parameters[3:])

            formula = f"=XLOOKUP({', '.join(map(str, formula_parts))})"
            self.active_sheet[command.target_range] = formula

    def _apply_filter(self, command: ExcelCommand) -> None:
        """
        FILTER 함수를 적용합니다.
        조건에 맞는 데이터만 필터링하여 반환합니다.

        Args:
            command: ExcelCommand 객체
                - parameters[0]: 필터링할 범위
                - parameters[1]: 조건
                - parameters[2]: (선택) 조건에 맞는 값이 없을 때 반환할 값
        """
        if command.parameters and len(command.parameters) >= 2:
            array = command.parameters[0]
            include = command.parameters[1]

            if len(command.parameters) >= 3:
                if_empty = command.parameters[2]
                formula = f"=FILTER({array}, {include}, {if_empty})"
            else:
                formula = f"=FILTER({array}, {include})"

            self.active_sheet[command.target_range] = formula

    def _apply_unique(self, command: ExcelCommand) -> None:
        """
        UNIQUE 함수를 적용합니다.
        중복을 제거한 고유값만 반환합니다.

        Args:
            command: ExcelCommand 객체
                - parameters[0]: 고유값을 추출할 범위
                - parameters[1]: (선택) by_col - True면 열 기준, False면 행 기준
                - parameters[2]: (선택) exactly_once - True면 정확히 한 번만 나타나는 값만 반환
        """
        if command.parameters and len(command.parameters) >= 1:
            array = command.parameters[0]

            # 기본 UNIQUE 수식
            formula_parts = [array]

            # 선택적 매개변수 추가
            if len(command.parameters) > 1:
                by_col = command.parameters[1] if len(command.parameters) > 1 else "FALSE"
                exactly_once = command.parameters[2] if len(command.parameters) > 2 else "FALSE"
                formula_parts.extend([by_col, exactly_once])

            formula = f"=UNIQUE({', '.join(map(str, formula_parts))})"
            self.active_sheet[command.target_range] = formula

    # 통계 함수 관련 메소드들
    def _apply_median(self, command: ExcelCommand) -> None:
        """
        MEDIAN 함수를 적용합니다.
        중간값(중위수)을 계산합니다.

        Args:
            command: ExcelCommand 객체
                - parameters[0]: 중위수를 계산할 범위
        """
        if command.parameters and "range" in command.parameters:
            range_str = command.parameters["range"]
            formula = f"=MEDIAN({range_str})"
            self.active_sheet[command.target_range] = formula

    def _apply_mode(self, command: ExcelCommand) -> None:
        """
        MODE 함수를 적용합니다.
        가장 자주 나타나는 값(최빈값)을 반환합니다.
        참고: Excel 2010 이후 MODE.SNGL 사용 권장

        Args:
            command: ExcelCommand 객체
                - parameters[0]: 최빈값을 계산할 범위
        """
        if command.parameters and "range" in command.parameters:
            range_str = command.parameters["range"]
            # MODE.SNGL 사용 (Excel 2010 이후 권장)
            formula = f"=MODE.SNGL({range_str})"
            self.active_sheet[command.target_range] = formula

    def _apply_stdev(self, command: ExcelCommand) -> None:
        """
        STDEV 함수를 적용합니다.
        표준편차를 계산합니다.
        참고: STDEV.S는 표본 표준편차, STDEV.P는 모집단 표준편차

        Args:
            command: ExcelCommand 객체
                - parameters[0]: 표준편차를 계산할 범위
                - parameters[1]: (선택) "S" 또는 "P" - 표본/모집단 구분
        """
        if command.parameters and "range" in command.parameters:
            range_str = command.parameters["range"]

            # 표본/모집단 구분 (기본값: 표본)
            stdev_type = command.parameters.get("type", "S")

            if stdev_type.upper() == "P":
                formula = f"=STDEV.P({range_str})"
            else:
                formula = f"=STDEV.S({range_str})"

            self.active_sheet[command.target_range] = formula

    def _apply_rank(self, command: ExcelCommand) -> None:
        """
        RANK 함수를 적용합니다.
        특정 값이 전체에서 몇 번째 순위인지 반환합니다.
        참고: RANK.EQ 사용 (Excel 2010 이후 권장)

        Args:
            command: ExcelCommand 객체
                - parameters[0]: 순위를 구할 값 또는 셀 참조
                - parameters[1]: 비교할 범위
                - parameters[2]: (선택) 순서 - 0 또는 생략: 내림차순, 1: 오름차순
        """
        if command.parameters and len(command.parameters) >= 2:
            number = command.parameters[0]
            ref = command.parameters[1]

            # 순서 매개변수 (기본값: 0 - 내림차순)
            order = command.parameters[2] if len(command.parameters) > 2 else "0"

            formula = f"=RANK.EQ({number}, {ref}, {order})"
            self.active_sheet[command.target_range] = formula

    def log_worksheet_contents(self, log_title: str = "워크시트 내용") -> None:
        """
        현재 워크시트의 모든 내용을 로그에 출력합니다.

        Args:
            log_title: 로그 제목
        """
        if not self.workbook or not self.active_sheet:
            print(f"[{log_title}] 워크북이 로드되지 않았습니다.")
            return

        print(f"\n{'=' * 50}")
        print(f"[{log_title}]")
        print(f"{'=' * 50}")

        # 워크시트 기본 정보
        ws = self.active_sheet
        max_row = ws.max_row
        max_col = ws.max_column

        print(f"워크시트명: {ws.title}")
        print(f"최대 행: {max_row}, 최대 열: {max_col}")

        if max_row == 1 and max_col == 1 and ws.cell(1, 1).value is None:
            print("워크시트가 비어있습니다.")
            print(f"{'=' * 50}\n")
            return

        # 헤더 출력 (열 번호)
        print("\n   ", end="")
        for col in range(1, max_col + 1):
            col_letter = ws.cell(1, col).column_letter
            print(f"{col_letter:>12}", end="")
        print()

        # 각 행의 데이터 출력
        for row in range(1, max_row + 1):
            print(f"{row:>3}:", end="")

            for col in range(1, max_col + 1):
                cell = ws.cell(row, col)

                # 셀 값 가져오기
                value = cell.value

                # 값이 None인 경우 공백 표시
                if value is None:
                    display_value = ""
                # 수식인 경우 수식과 결과값 모두 표시
                elif isinstance(value, str) and value.startswith('='):
                    try:
                        # 수식의 계산 결과 시도
                        calculated_value = cell.displayed_value if hasattr(cell, 'displayed_value') else "계산필요"
                        display_value = f"{value}({calculated_value})"
                    except:
                        display_value = value
                else:
                    display_value = str(value)

                # 너무 긴 값은 잘라내기
                if len(display_value) > 10:
                    display_value = display_value[:7] + "..."

                print(f"{display_value:>12}", end="")
            print()  # 행 끝에서 줄바꿈

        # 수식 정보 별도 출력
        print(f"\n[수식 정보]")
        formula_found = False
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                cell = ws.cell(row, col)
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    print(f"  {cell.coordinate}: {cell.value}")
                    formula_found = True

        if not formula_found:
            print("  수식이 없습니다.")

        # 서식 정보 출력 (굵게, 색상 등)
        print(f"\n[서식 정보]")
        formatted_cells = []
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                cell = ws.cell(row, col)
                format_info = []

                # 폰트 정보
                if cell.font:
                    if cell.font.bold:
                        format_info.append("굵게")
                    if cell.font.italic:
                        format_info.append("기울임")
                    if cell.font.underline:
                        format_info.append("밑줄")

                    # 글자색 정보 (안전하게 처리)
                    if cell.font.color:
                        try:
                            if hasattr(cell.font.color, 'rgb') and cell.font.color.rgb:
                                # rgb 값이 문자열인지 확인
                                rgb_value = str(cell.font.color.rgb)
                                if rgb_value != "FF000000" and rgb_value != "None":
                                    format_info.append(f"글자색:{rgb_value}")
                            elif hasattr(cell.font.color, 'theme') and cell.font.color.theme is not None:
                                format_info.append(f"글자색:테마{cell.font.color.theme}")
                            elif hasattr(cell.font.color, 'indexed') and cell.font.color.indexed is not None:
                                format_info.append(f"글자색:인덱스{cell.font.color.indexed}")
                        except Exception as e:
                            # 색상 정보를 가져오는 데 실패한 경우
                            pass

                # 배경색 정보 (안전하게 처리)
                if cell.fill and cell.fill.start_color:
                    try:
                        if hasattr(cell.fill.start_color, 'rgb') and cell.fill.start_color.rgb:
                            rgb_value = str(cell.fill.start_color.rgb)
                            if rgb_value != "00000000" and rgb_value != "None":
                                format_info.append(f"배경색:{rgb_value}")
                        elif hasattr(cell.fill.start_color, 'theme') and cell.fill.start_color.theme is not None:
                            format_info.append(f"배경색:테마{cell.fill.start_color.theme}")
                        elif hasattr(cell.fill.start_color, 'indexed') and cell.fill.start_color.indexed is not None:
                            format_info.append(f"배경색:인덱스{cell.fill.start_color.indexed}")
                    except Exception as e:
                        # 배경색 정보를 가져오는 데 실패한 경우
                        pass

                # 정렬 정보
                if cell.alignment:
                    if cell.alignment.horizontal and cell.alignment.horizontal != "general":
                        format_info.append(f"수평:{cell.alignment.horizontal}")
                    if cell.alignment.vertical and cell.alignment.vertical != "bottom":
                        format_info.append(f"수직:{cell.alignment.vertical}")

                if format_info:
                    formatted_cells.append(f"  {cell.coordinate}: {', '.join(format_info)}")

        if formatted_cells:
            for info in formatted_cells:
                print(info)
        else:
            print("  서식이 적용된 셀이 없습니다.")

        print(f"{'=' * 50}\n")


# process_excel_with_commands 함수 수정
def process_excel_with_commands(
        excel_bytes: bytes,
        commands: Any
) -> bytes:
    """
    엑셀 파일에 명령어를 적용하고 결과를 반환합니다.

    Args:
        excel_bytes: 원본 엑셀 파일의 바이트 데이터
        commands: 적용할 명령어 리스트

    Returns:
        수정된 엑셀 파일의 바이트 데이터
    """
    manipulator = ExcelManipulator()

    # 엑셀 파일 로드
    manipulator.load_from_bytes(excel_bytes)

    # 🔹 수정 전 상태 로그 출력
    manipulator.log_worksheet_contents("명령어 적용 전 워크시트 상태")

    # 명령어 실행
    print(f"\n[실행할 명령어 목록]")
    for i, command in enumerate(commands, 1):
        print(f"  {i}. {command.command_type} -> {command.target_range} | {command.parameters}")
    print()

    manipulator.execute_commands(commands)

    # 🔹 수정 후 상태 로그 출력
    manipulator.log_worksheet_contents("명령어 적용 후 워크시트 최종 상태")

    # 결과 저장 및 반환
    return manipulator.save_to_bytes()


def create_empty_excel() -> bytes:
    """
    빈 엑셀 파일을 생성합니다.

    Returns:
        빈 엑셀 파일의 바이트 데이터
    """
    workbook = Workbook()
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()

