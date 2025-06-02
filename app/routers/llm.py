# app/routers/llm_test.py
"""
LLM 서비스 테스트를 위한 라우터
엑셀 데이터를 JSON 형태로 주고받으며 LLM 기능을 테스트할 수 있습니다.
"""

from fastapi import APIRouter, HTTPException, status
from pydantic import BaseModel
from typing import List, Dict, Any, Optional
import json
import io
from openpyxl import Workbook

from app.services.llm import get_llm_response
from app.services.excel import process_excel_with_commands, create_empty_excel
from app.schemas.excel import ExcelCommand
from typing import List

router = APIRouter()

class CommandSequenceTestRequest(BaseModel):
    """명령어 시퀀스 테스트 요청 스키마"""
    commands: List[Dict[str, Any]]  # 실행할 명령어 시퀀스
    initial_data: Optional[List[List[Any]]] = None  # 초기 데이터 (선택사항)

class CommandSequenceTestResponse(BaseModel):
    """명령어 시퀀스 테스트 응답 스키마"""
    success: bool  # 성공 여부
    message: str  # 결과 메시지
    initial_data: List[List[Any]]  # 초기 데이터
    final_data: List[List[Any]]  # 최종 데이터
    executed_commands: List[Dict[str, Any]]  # 실행된 명령어들
    errors: List[str]  # 오류 메시지들


# 요청/응답 스키마 정의
class ExcelDataRequest(BaseModel):
    """엑셀 데이터 요청 스키마"""
    data: List[List[Any]]  # 2차원 배열 형태의 엑셀 데이터
    summary: Optional[str] = ""  # 이전 대화 요약
    command: str  # 사용자 명령


class ExcelDataResponse(BaseModel):
    """엑셀 데이터 응답 스키마"""
    message: str  # LLM 응답 메시지
    data: List[List[Any]]  # 수정된 엑셀 데이터
    commands: List[Dict[str, Any]]  # 실행된 명령어들
    summary: str  # 업데이트된 요약


class CommandTestRequest(BaseModel):
    """명령어 테스트 요청 스키마"""
    command: str  # 자연어 명령
    summary: Optional[str] = ""  # 대화 요약
    excel_context: Optional[str] = ""  # 엑셀 컨텍스트


class CommandTestResponse(BaseModel):
    """명령어 테스트 응답 스키마"""
    chat: str  # LLM 응답
    commands: List[Dict[str, Any]]  # 생성된 명령어들
    summary: str  # 업데이트된 요약


class ExcelSimulationRequest(BaseModel):
    """엑셀 시뮬레이션 요청 스키마"""
    initial_data: List[List[Any]]  # 초기 엑셀 데이터
    commands_sequence: List[str]  # 연속된 명령어들


class ExcelSimulationResponse(BaseModel):
    """엑셀 시뮬레이션 응답 스키마"""
    steps: List[Dict[str, Any]]  # 각 단계별 결과
    final_data: List[List[Any]]  # 최종 엑셀 데이터
    summary: str  # 최종 요약


@router.post(
    "/test/basic",
    response_model=ExcelDataResponse,
    summary="기본 LLM 엑셀 처리 테스트",
    description="JSON 형태의 엑셀 데이터를 받아 자연어 명령을 처리하고 결과를 반환합니다."
)
async def test_basic_llm_excel(request: ExcelDataRequest):
    """
    기본적인 LLM 엑셀 처리 기능을 테스트합니다.

    이 엔드포인트는 다음을 테스트합니다:
    - JSON 데이터를 엑셀 바이트로 변환
    - LLM 서비스 호출
    - 엑셀 명령어 실행
    - 결과를 JSON으로 변환
    """
    try:
        # 1. JSON 데이터를 엑셀 바이트로 변환
        excel_bytes = _json_to_excel_bytes(request.data)

        # 2. LLM 서비스 호출
        llm_result = get_llm_response(
            user_command=request.command,
            excel_bytes=excel_bytes,
            session_summary=request.summary
        )

        # 3. 엑셀 명령어 실행
        modified_excel_bytes = process_excel_with_commands(
            excel_bytes=excel_bytes,
            commands=llm_result.cmd_seq
        )

        # 4. 결과를 JSON으로 변환
        result_data = _excel_bytes_to_json(modified_excel_bytes)

        # 5. 명령어들을 딕셔너리로 변환
        commands_dict = [
            {
                "command_type": cmd.command_type,
                "target_range": cmd.target_range,
                "parameters": cmd.parameters
            }
            for cmd in llm_result.cmd_seq
        ]

        return ExcelDataResponse(
            message=llm_result.chat,
            data=result_data,
            commands=commands_dict,
            summary=llm_result.summary
        )

    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"LLM 처리 중 오류 발생: {str(e)}"
        )


@router.post(
    "/test/command-only",
    response_model=CommandTestResponse,
    summary="명령어 생성 테스트",
    description="엑셀 파일 없이 자연어 명령만으로 명령어 시퀀스 생성을 테스트합니다."
)
async def test_command_generation(request: CommandTestRequest):
    """
    명령어 생성 기능만 테스트합니다.
    실제 엑셀 파일 처리 없이 LLM이 어떤 명령어를 생성하는지 확인할 수 있습니다.
    """
    try:
        # 빈 엑셀 파일 생성
        empty_excel_bytes = create_empty_excel()

        # LLM 서비스 호출 (명령어 생성만)
        llm_result = get_llm_response(
            user_command=request.command,
            excel_bytes=empty_excel_bytes,
            session_summary=request.summary
        )

        # 명령어들을 딕셔너리로 변환
        commands_dict = [
            {
                "command_type": cmd.command_type,
                "target_range": cmd.target_range,
                "parameters": cmd.parameters
            }
            for cmd in llm_result.cmd_seq
        ]

        return CommandTestResponse(
            chat=llm_result.chat,
            commands=commands_dict,
            summary=llm_result.summary
        )

    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"명령어 생성 중 오류 발생: {str(e)}"
        )


@router.post(
    "/test/simulation",
    response_model=ExcelSimulationResponse,
    summary="연속 명령 시뮬레이션 테스트",
    description="여러 명령어를 연속으로 실행하여 전체 플랫폼 동작을 시뮬레이션합니다."
)
async def test_excel_simulation(request: ExcelSimulationRequest):
    """
    여러 명령어를 연속으로 실행하는 시뮬레이션을 수행합니다.
    실제 채팅 세션과 유사한 환경에서 테스트할 수 있습니다.
    """
    try:
        # 초기 데이터 설정
        current_data = request.initial_data
        current_summary = ""
        steps = []

        # 각 명령어를 순차적으로 실행
        for i, command in enumerate(request.commands_sequence):
            # JSON 데이터를 엑셀 바이트로 변환
            excel_bytes = _json_to_excel_bytes(current_data)

            # LLM 서비스 호출
            llm_result = get_llm_response(
                user_command=command,
                excel_bytes=excel_bytes,
                session_summary=current_summary
            )

            # 엑셀 명령어 실행
            modified_excel_bytes = process_excel_with_commands(
                excel_bytes=excel_bytes,
                commands=llm_result.cmd_seq
            )

            # 결과를 JSON으로 변환
            current_data = _excel_bytes_to_json(modified_excel_bytes)
            current_summary = llm_result.summary

            # 단계별 결과 저장
            step_result = {
                "step": i + 1,
                "command": command,
                "response": llm_result.chat,
                "executed_commands": [
                    {
                        "command_type": cmd.command_type,
                        "target_range": cmd.target_range,
                        "parameters": cmd.parameters
                    }
                    for cmd in llm_result.cmd_seq
                ],
                "data_after": current_data.copy()
            }
            steps.append(step_result)

        return ExcelSimulationResponse(
            steps=steps,
            final_data=current_data,
            summary=current_summary
        )

    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"시뮬레이션 중 오류 발생: {str(e)}"
        )


@router.get(
    "/test/examples",
    summary="테스트 예시 데이터 제공",
    description="다양한 테스트 시나리오를 위한 예시 데이터를 제공합니다."
)
async def get_test_examples():
    """
    테스트를 위한 예시 데이터들을 제공합니다.
    """
    examples = {
        "basic_data": [
            ["이름", "나이", "점수"],
            ["김철수", 25, 85],
            ["이영희", 23, 92],
            ["박민수", 27, 78],
            ["정수진", 24, 88]
        ],
        "financial_data": [
            ["항목", "1월", "2월", "3월"],
            ["매출", 1000000, 1200000, 1100000],
            ["비용", 800000, 850000, 900000],
            ["순이익", "", "", ""],  # 계산이 필요한 셀
        ],
        "inventory_data": [
            ["상품명", "재고수량", "단가", "총액"],
            ["노트북", 10, 1200000, ""],
            ["마우스", 50, 25000, ""],
            ["키보드", 30, 75000, ""],
        ],
        "sample_commands": [
            "점수 열의 평균을 구해서 하단에 표시해주세요",
            "총액 열을 계산해서 채워주세요 (재고수량 × 단가)",
            "헤더 행을 굵게 만들어주세요",
            "점수가 80점 이상인 셀들을 파란색으로 칠해주세요",
            "전체 데이터에 테두리를 추가해주세요"
        ],
        "simulation_scenario": {
            "initial_data": [
                ["제품", "판매량", "단가", "매출"],
                ["A제품", 100, 10000, ""],
                ["B제품", 150, 15000, ""],
                ["C제품", 80, 12000, ""]
            ],
            "commands": [
                "매출 열을 계산해주세요 (판매량 × 단가)",
                "전체 매출의 합계를 구해서 하단에 표시해주세요",
                "헤더를 굵게 만들고 배경색을 연한 파란색으로 해주세요",
                "매출이 가장 높은 제품을 빨간색으로 강조해주세요"
            ]
        }
    }

    return examples


# 헬퍼 함수들
def _json_to_excel_bytes(data: List[List[Any]]) -> bytes:
    """
    JSON 배열 데이터를 엑셀 바이트로 변환합니다.

    Args:
        data: 2차원 배열 형태의 데이터

    Returns:
        엑셀 파일의 바이트 데이터
    """
    workbook = Workbook()
    worksheet = workbook.active

    # 데이터를 엑셀 시트에 쓰기
    for row_idx, row_data in enumerate(data, 1):
        for col_idx, cell_value in enumerate(row_data, 1):
            worksheet.cell(row=row_idx, column=col_idx, value=cell_value)

    # 바이트로 변환
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()


def _excel_bytes_to_json(excel_bytes: bytes) -> List[List[Any]]:
    """
    엑셀 바이트 데이터를 JSON 배열로 변환합니다.

    Args:
        excel_bytes: 엑셀 파일의 바이트 데이터

    Returns:
        2차원 배열 형태의 데이터
    """
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(excel_bytes))
    worksheet = workbook.active

    # 데이터가 있는 범위 확인
    max_row = worksheet.max_row
    max_col = worksheet.max_column

    # 데이터 추출
    data = []
    for row in range(1, max_row + 1):
        row_data = []
        for col in range(1, max_col + 1):
            cell_value = worksheet.cell(row=row, column=col).value
            # None 값을 빈 문자열로 변환
            row_data.append(cell_value if cell_value is not None else "")
        data.append(row_data)

    return data


# 기존 엔드포인트들 아래에 새로운 엔드포인트 추가
@router.post(
    "/test/command-sequence",
    response_model=CommandSequenceTestResponse,
    summary="명령어 시퀀스 직접 테스트",
    description="JSON 형태의 명령어 시퀀스를 직접 입력받아 실행하고 결과를 확인합니다."
)
async def test_command_sequence(request: CommandSequenceTestRequest):
    """
    명령어 시퀀스를 직접 테스트합니다.

    이 엔드포인트는 다음을 수행합니다:
    1. 빈 엑셀 파일 또는 제공된 초기 데이터로 엑셀 생성
    2. 각 명령어를 순차적으로 실행
    3. 각 단계에서 발생한 오류 수집
    4. 최종 결과를 JSON으로 반환

    Example request:
    ```json
    {
        "commands": [
            {
                "command_type": "set_value",
                "target_range": "A1",
                "parameters": {"value": "이름"}
            },
            {
                "command_type": "set_value",
                "target_range": "B1",
                "parameters": {"value": "점수"}
            },
            {
                "command_type": "bold",
                "target_range": "A1:B1",
                "parameters": {}
            }
        ]
    }
    ```
    """
    errors = []
    executed_commands = []

    try:
        # 1. 초기 엑셀 파일 생성
        if request.initial_data:
            # 제공된 초기 데이터로 엑셀 생성
            excel_bytes = _json_to_excel_bytes(request.initial_data)
            initial_data = request.initial_data
        else:
            # 빈 엑셀 파일 생성
            excel_bytes = create_empty_excel()
            initial_data = [[]]  # 빈 데이터

        # 2. 명령어 시퀀스를 ExcelCommand 객체로 변환
        excel_commands = []
        for idx, cmd in enumerate(request.commands):
            try:
                # 명령어 검증 및 변환
                if "command_type" not in cmd:
                    errors.append(f"명령어 {idx + 1}: command_type이 없습니다.")
                    continue

                if "target_range" not in cmd:
                    errors.append(f"명령어 {idx + 1}: target_range가 없습니다.")
                    continue

                # parameters가 없으면 빈 딕셔너리로 설정
                parameters = cmd.get("parameters", {})

                excel_command = ExcelCommand(
                    command_type=cmd["command_type"],
                    target_range=cmd["target_range"],
                    parameters=parameters
                )
                excel_commands.append(excel_command)

                # 실행된 명령어 기록
                executed_commands.append({
                    "index": idx + 1,
                    "command_type": excel_command.command_type,
                    "target_range": excel_command.target_range,
                    "parameters": excel_command.parameters,
                    "status": "prepared"
                })

            except Exception as e:
                errors.append(f"명령어 {idx + 1} 변환 오류: {str(e)}")
                executed_commands.append({
                    "index": idx + 1,
                    "command": cmd,
                    "status": "conversion_failed",
                    "error": str(e)
                })

        # 3. 명령어 실행
        if excel_commands:
            try:
                # process_excel_with_commands 함수 사용
                modified_excel_bytes = process_excel_with_commands(
                    excel_bytes=excel_bytes,
                    commands=excel_commands
                )

                # 실행 상태 업데이트
                for cmd in executed_commands:
                    if cmd.get("status") == "prepared":
                        cmd["status"] = "executed"

            except Exception as e:
                errors.append(f"명령어 실행 중 오류: {str(e)}")
                modified_excel_bytes = excel_bytes  # 오류 발생시 원본 유지
        else:
            modified_excel_bytes = excel_bytes
            errors.append("실행할 유효한 명령어가 없습니다.")

        # 4. 결과를 JSON으로 변환
        final_data = _excel_bytes_to_json(modified_excel_bytes)

        # 5. 응답 생성
        success = len(errors) == 0
        message = "명령어 시퀀스가 성공적으로 실행되었습니다." if success else f"{len(errors)}개의 오류가 발생했습니다."

        return CommandSequenceTestResponse(
            success=success,
            message=message,
            initial_data=initial_data,
            final_data=final_data,
            executed_commands=executed_commands,
            errors=errors
        )

    except Exception as e:
        return CommandSequenceTestResponse(
            success=False,
            message=f"처리 중 오류 발생: {str(e)}",
            initial_data=[[]],
            final_data=[[]],
            executed_commands=executed_commands,
            errors=[str(e)]
        )


@router.post(
    "/test/validate-commands",
    summary="명령어 유효성 검증",
    description="명령어 시퀀스의 유효성을 검증합니다 (실행하지 않음)."
)
async def validate_commands(request: CommandSequenceTestRequest):
    """
    명령어 시퀀스의 유효성만 검증합니다.
    실제로 실행하지는 않고, 각 명령어가 올바른 형식인지 확인합니다.
    """
    validation_results = []
    valid_command_types = [
        # 함수 관련
        "sum", "average", "count", "max", "min",
        # 서식 관련
        "bold", "italic", "underline",
        "font_color", "fill_color", "border",
        "font_size", "font_name",
        # 데이터 관련
        "set_value", "clear", "merge", "unmerge",
        # 정렬 관련
        "align_left", "align_center", "align_right",
        "align_top", "align_middle", "align_bottom"
    ]

    for idx, cmd in enumerate(request.commands):
        result = {
            "index": idx + 1,
            "command": cmd,
            "valid": True,
            "errors": []
        }

        # command_type 검증
        if "command_type" not in cmd:
            result["valid"] = False
            result["errors"].append("command_type이 없습니다.")
        elif cmd["command_type"] not in valid_command_types:
            result["valid"] = False
            result["errors"].append(f"유효하지 않은 command_type: {cmd['command_type']}")

        # target_range 검증
        if "target_range" not in cmd:
            result["valid"] = False
            result["errors"].append("target_range가 없습니다.")
        else:
            # 셀 범위 형식 검증 (간단한 정규식)
            import re
            cell_pattern = r'^[A-Z]+\d+(?::[A-Z]+\d+)?$'
            if not re.match(cell_pattern, cmd["target_range"]):
                result["valid"] = False
                result["errors"].append(f"유효하지 않은 target_range 형식: {cmd['target_range']}")

        # parameters 검증
        if "parameters" in cmd:
            cmd_type = cmd.get("command_type", "")
            params = cmd["parameters"]

            # 명령어별 파라미터 검증
            if cmd_type in ["sum", "average", "count", "max", "min"]:
                if not params.get("range"):
                    result["errors"].append("수식 명령어는 'range' 파라미터가 필요합니다.")
                    result["valid"] = False
            elif cmd_type in ["font_color", "fill_color"]:
                if not params.get("color"):
                    result["errors"].append("색상 명령어는 'color' 파라미터가 필요합니다.")
                    result["valid"] = False
                elif not re.match(r'^[0-9A-Fa-f]{6}$', params.get("color", "")):
                    result["errors"].append("색상은 6자리 16진수여야 합니다 (예: FF0000).")
                    result["valid"] = False
            elif cmd_type == "set_value":
                if "value" not in params:
                    result["errors"].append("set_value 명령어는 'value' 파라미터가 필요합니다.")
                    result["valid"] = False

        validation_results.append(result)

    # 전체 요약
    total_commands = len(request.commands)
    valid_commands = sum(1 for r in validation_results if r["valid"])

    return {
        "summary": {
            "total_commands": total_commands,
            "valid_commands": valid_commands,
            "invalid_commands": total_commands - valid_commands,
            "all_valid": valid_commands == total_commands
        },
        "validation_results": validation_results
    }